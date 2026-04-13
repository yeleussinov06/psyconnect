"""
python manage.py import_users file.xlsx
python manage.py import_users file.xlsx --dry-run
python manage.py import_users file.xlsx --output result.xlsx
"""
import os, re, secrets, string
from django.core.management.base import BaseCommand
from django.contrib.auth.models import User
from django.db import transaction
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def next_student_id():
    """Генерирует следующий авто-ID: id0001, id0002..."""
    existing = User.objects.filter(username__regex=r'^id\d+$') \
        .values_list('username', flat=True)
    nums = []
    for u in existing:
        try:
            nums.append(int(u[2:]))
        except ValueError:
            pass
    next_num = max(nums) + 1 if nums else 1
    return f'id{next_num:04d}'


def gen_password(base=''):
    if base:
        return f'{base}_2026'
    chars = string.ascii_letters + string.digits
    return ''.join(secrets.choice(chars) for _ in range(10))


def make_username(full_name, prefix=''):
    """Генерирует username из ФИО"""
    parts = re.sub(r'[^\w\s]', '', full_name.lower()).split()
    base = prefix + '_'.join(parts[:2]) if parts else prefix + 'user'
    base = re.sub(r'[^a-z0-9_а-яё]', '_', base)
    uname = base
    counter = 1
    while User.objects.filter(username=uname).exists():
        uname = f'{base}_{counter}'
        counter += 1
    return uname


def find_or_create_teacher(full_name, class_obj):
    """Находит или создаёт учителя и привязывает к классу"""
    from core.models import UserProfile
    # Ищем по имени среди учителей
    existing = User.objects.filter(
        profile__role='teacher',
        first_name=full_name.split()[0] if full_name.split() else '',
    )
    if full_name.split().__len__() > 1:
        existing = existing.filter(last_name__icontains=full_name.split()[1])

    if existing.exists():
        teacher = existing.first()
        # Привязываем к классу если ещё не привязан
        profile = teacher.profile
        if profile.school_class != class_obj:
            profile.school_class = class_obj
            profile.save()
        return teacher, False  # False = не новый

    # Создаём нового учителя
    uname = make_username(full_name)
    pwd = gen_password()
    parts = full_name.split()
    teacher = User.objects.create_user(
        username=uname,
        password=pwd,
        first_name=parts[0] if parts else '',
        last_name=' '.join(parts[1:]) if len(parts) > 1 else '',
    )
    profile, _ = UserProfile.objects.get_or_create(user=teacher)
    profile.role = 'teacher'
    profile.school_class = class_obj
    profile.save()
    return teacher, pwd  # pwd = новый


def find_or_create_parent(full_name, phone, email=''):
    """Находит родителя по телефону или создаёт нового"""
    from core.models import UserProfile
    # Handle numeric phone (Excel stores as int)
    if isinstance(phone, (int, float)):
        phone = str(int(phone))
    phone_clean = re.sub(r'[\s\-\(\)\+]', '', str(phone)) if phone else ''

    # Ищем по телефону
    if phone_clean:
        existing = UserProfile.objects.filter(
            role='parent', phone__icontains=phone_clean[-7:]  # последние 7 цифр
        ).select_related('user')
        if existing.exists():
            return existing.first().user, False  # уже существует

    # Ищем по email
    if email:
        existing = User.objects.filter(email=email, profile__role='parent')
        if existing.exists():
            return existing.first(), False

    # Создаём нового родителя
    uname = make_username(full_name, prefix='parent_')
    pwd = gen_password()
    parts = full_name.split()
    parent = User.objects.create_user(
        username=uname,
        email=email or '',
        password=pwd,
        first_name=parts[0] if parts else '',
        last_name=' '.join(parts[1:]) if len(parts) > 1 else '',
    )
    profile, _ = UserProfile.objects.get_or_create(user=parent)
    profile.role = 'parent'
    profile.phone = phone or ''
    profile.save()
    return parent, pwd  # pwd = новый


class Command(BaseCommand):
    help = 'Импорт класса из Excel (один лист)'

    def add_arguments(self, parser):
        parser.add_argument('file', type=str)
        parser.add_argument('--output', type=str, default='import_result.xlsx')
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **options):
        file_path = options['file']
        output_path = options['output']
        dry_run = options['dry_run']

        if not os.path.exists(file_path):
            self.stderr.write(self.style.ERROR(f'Файл не найден: {file_path}'))
            return

        self.stdout.write(self.style.MIGRATE_HEADING(
            f'\n{"[DRY RUN] " if dry_run else ""}Импорт из: {file_path}\n'
        ))

        try:
            wb = load_workbook(file_path)
        except Exception as e:
            self.stderr.write(self.style.ERROR(f'Ошибка чтения: {e}'))
            return

        # Берём первый лист (или "Импорт класса")
        ws = wb['Импорт класса'] if 'Импорт класса' in wb.sheetnames else wb.active

        stats = {'teachers': 0, 'students': 0, 'parents': 0, 'skipped': 0}
        errors = []
        created_users = []

        try:
            with transaction.atomic():
                sid = transaction.savepoint()

                from core.models import UserProfile, SchoolClass, ParentStudent

                # Кэш учителей и классов в рамках импорта
                teacher_cache = {}  # full_name → (user, pwd)
                class_cache = {}    # name → SchoolClass

                # Пропускаем строки 1-5 (заголовок + примеры + примечание)
                for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                    if not any(row):
                        continue

                    try:
                        class_name    = str(row[0] or '').strip()
                        student_name  = str(row[1] or '').strip()
                        student_email = str(row[2] or '').strip() if len(row) > 2 else ''
                        teacher_name  = str(row[3] or '').strip() if len(row) > 3 else ''
                        parent_name   = str(row[4] or '').strip() if len(row) > 4 else ''
                        parent_phone  = str(row[5] or '').strip() if len(row) > 5 else ''
                        parent_email  = str(row[6] or '').strip() if len(row) > 6 else ''

                        if not class_name or not student_name:
                            continue

                        # ── Класс ──────────────────────────────
                        if class_name not in class_cache:
                            sc, _ = SchoolClass.objects.get_or_create(name=class_name)
                            class_cache[class_name] = sc
                        school_class = class_cache[class_name]

                        # ── Учитель ────────────────────────────
                        teacher_pwd = None
                        if teacher_name and teacher_name not in teacher_cache:
                            teacher, pwd = find_or_create_teacher(teacher_name, school_class)
                            teacher_cache[teacher_name] = teacher
                            if pwd:  # новый
                                stats['teachers'] += 1
                                teacher_pwd = pwd
                                created_users.append({
                                    'role': 'Учитель',
                                    'full_name': teacher_name,
                                    'username': teacher.username,
                                    'password': pwd,
                                    'email': teacher.email,
                                    'class': class_name,
                                })
                                self.stdout.write(f'  Учитель: {teacher_name} ({teacher.username})')

                        # ── Ученик ─────────────────────────────
                        # Проверяем дубль по имени в том же классе
                        existing_student = User.objects.filter(
                            profile__role='student',
                            profile__school_class=school_class,
                            first_name=student_name.split()[0] if student_name.split() else '',
                        )
                        if student_name.split().__len__() > 1:
                            existing_student = existing_student.filter(
                                last_name__icontains=student_name.split()[1]
                            )

                        if existing_student.exists():
                            student = existing_student.first()
                            self.stdout.write(f'  Пропуск (уже есть): {student_name}')
                            stats['skipped'] += 1
                        else:
                            student_id = next_student_id()
                            pwd = gen_password(student_id)
                            parts = student_name.split()
                            student = User.objects.create_user(
                                username=student_id,
                                email=student_email,
                                password=pwd,
                                first_name=parts[0] if parts else '',
                                last_name=' '.join(parts[1:]) if len(parts) > 1 else '',
                            )
                            profile, _ = UserProfile.objects.get_or_create(user=student)
                            profile.role = 'student'
                            profile.school_class = school_class
                            profile.save()
                            stats['students'] += 1
                            created_users.append({
                                'role': 'Ученик',
                                'full_name': student_name,
                                'username': student_id,
                                'password': pwd,
                                'email': student_email,
                                'class': class_name,
                                'student_id': student_id,
                            })
                            self.stdout.write(
                                self.style.SUCCESS(f'  Ученик: {student_name} → {student_id}')
                            )

                        # ── Родитель ───────────────────────────
                        if parent_name and parent_phone:
                            parent, pwd = find_or_create_parent(
                                parent_name, parent_phone, parent_email
                            )
                            # Привязываем ребёнка
                            link, created = ParentStudent.objects.get_or_create(
                                parent=parent, student=student
                            )
                            if pwd:  # новый родитель
                                stats['parents'] += 1
                                created_users.append({
                                    'role': 'Родитель',
                                    'full_name': parent_name,
                                    'username': parent.username,
                                    'password': pwd,
                                    'email': parent_email,
                                    'phone': parent_phone,
                                    'children': student.username,
                                })
                                self.stdout.write(f'  Родитель: {parent_name} ({parent.username})')
                            else:
                                if created:
                                    self.stdout.write(
                                        f'  Родитель {parent_name}: привязан ещё один ребёнок'
                                    )

                    except Exception as e:
                        errors.append(f'Строка {row_num}: {e}')
                        self.stderr.write(self.style.ERROR(f'  ✗ Строка {row_num}: {e}'))

                if dry_run:
                    transaction.savepoint_rollback(sid)
                    self.stdout.write(self.style.WARNING('\n[DRY RUN] Откат. Пользователи не созданы.'))

        except Exception as e:
            self.stderr.write(self.style.ERROR(f'\nКритическая ошибка: {e}'))
            return

        # ── Отчёт ───────────────────────────────────────────────
        self.stdout.write('\n' + '─' * 55)
        self.stdout.write(self.style.SUCCESS(
            f'Создано: учителей={stats["teachers"]}, '
            f'учеников={stats["students"]}, '
            f'родителей={stats["parents"]}. '
            f'Пропущено={stats["skipped"]}.'
        ))
        if errors:
            self.stdout.write(self.style.ERROR(f'Ошибки ({len(errors)}):'))
            for err in errors:
                self.stdout.write(f'  • {err}')

        # ── Выходной Excel ───────────────────────────────────────
        if created_users and not dry_run:
            self._write_result(created_users, output_path)
            self.stdout.write(self.style.SUCCESS(f'Файл с паролями: {output_path}'))

    def _write_result(self, users, path):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Логины и пароли'

        h_fill = PatternFill('solid', fgColor='4A7C59')
        headers = ['Роль','ФИО','Логин','Пароль','Email','Телефон','Класс','ID']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True, color='FFFFFF', name='Calibri')
            cell.fill = h_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        colors = {'Учитель':'EBF3EE','Ученик':'EBF4FA','Родитель':'FDF3E8'}
        for r, u in enumerate(users, 2):
            fill = PatternFill('solid', fgColor=colors.get(u.get('role',''),'FFFFFF'))
            row_data = [
                u.get('role',''), u.get('full_name',''), u.get('username',''),
                u.get('password',''), u.get('email',''), u.get('phone',''),
                u.get('class',''), u.get('student_id', u.get('children','')),
            ]
            for c, v in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.fill = fill
                cell.font = Font(name='Calibri', size=10)

        widths = [12, 30, 18, 18, 28, 18, 10, 14]
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = 'A2'
        wb.save(path)
