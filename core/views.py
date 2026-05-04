from django.shortcuts import render, redirect, get_object_or_404
from django.db import models
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import JsonResponse
from django.views.decorators.http import require_POST
import json, random
from datetime import timedelta
from django.utils import timezone
from collections import Counter
from .models import UserProfile, EmotionEntry, TestResult, ChatMessage, AnonymousRequest, Article, PsychologistSchedule, Appointment, Notification, StudentRequest, TeacherObservation
from .models import PsychTest, PsychTestResult


_TONE_KEYWORDS = {
    'fear': [
        'угрожают', 'бьют', 'бьет', 'бьёт', 'буллинг', 'издеваются', 'обижают',
        'запугивают', 'боюсь идти', 'не хочу в школу', 'унижают',
        'смеются надо мной', 'дразнят', 'избивают', 'насилие', 'угрозы',
    ],
    'aggression': [
        'ненавижу', 'бесит', 'достал', 'достала', 'хочу ударить', 'злюсь',
        'бешусь', 'злость', 'ненависть', 'раздражает', 'злой', 'орёт',
        'орет', 'кричит', 'кричат', 'взрываюсь',
    ],
    'depression': [
        'нет смысла', 'не хочу жить', 'ничего не хочу', 'всё плохо',
        'нет сил', 'не вижу смысла', 'подавлен', 'подавлена', 'апатия',
        'безнадёжно', 'безнадежно', 'одиноко', 'никто не понимает',
        'никому не нужен', 'никому не нужна', 'грустно', 'плачу',
        'слёзы', 'слезы', 'устал от всего', 'устала от всего',
    ],
    'anxiety': [
        'боюсь', 'страшно', 'тревога', 'беспокоит', 'не могу спать',
        'переживаю', 'волнуюсь', 'нервничаю', 'стресс', 'паника',
        'тревожно', 'беспокойство', 'не уверен', 'не уверена',
        'пугает', 'сердце колотится', 'трясёт', 'трясет', 'страх',
    ],
    'confusion': [
        'не понимаю', 'запутался', 'запуталась', 'не знаю что делать',
        'растерян', 'растеряна', 'непонятно', 'что делать',
        'потерялся', 'потерялась', 'не знаю как',
    ],
}

_RISK_ORDER = ['low', 'medium', 'high', 'critical']

def _analyze_complaint(req):
    """
    Шаг 2: keyword-анализ текста жалобы.
    Заполняет ai_* поля StudentRequest и переводит статус в ai_analyzed.
    Вызывается синхронно сразу после создания StudentRequest.
    """
    text_lower = req.text.lower()
    student    = req.student
    now        = timezone.now()

    # ── 1. Классификация тональности ─────────────────────────
    tone_scores = {tone: 0 for tone in _TONE_KEYWORDS}
    matched_kw  = []
    for tone, kws in _TONE_KEYWORDS.items():
        for kw in kws:
            if kw in text_lower:
                tone_scores[tone] += 1
                matched_kw.append(kw)

    best_tone     = max(tone_scores, key=tone_scores.get)
    detected_tone = best_tone if tone_scores[best_tone] > 0 else 'neutral'

    # ── 2. Базовый риск по тону ───────────────────────────────
    base_risk = {'fear': 'high', 'aggression': 'medium', 'depression': 'medium',
                 'anxiety': 'low', 'confusion': 'low', 'neutral': 'low'}
    risk_idx = _RISK_ORDER.index(base_risk[detected_tone])

    # ── 3. Агрегация истории ──────────────────────────────────
    ago30 = now - timedelta(days=30)
    ago14 = now - timedelta(days=14)

    past_qs    = StudentRequest.objects.filter(student=student).exclude(id=req.id)
    past_total = past_qs.count()
    past_30d   = past_qs.filter(created_at__gte=ago30).count()

    obs_qs       = TeacherObservation.objects.filter(student=student)
    obs_total    = obs_qs.count()
    obs_30d      = obs_qs.filter(created_at__gte=ago30).count()
    obs_critical = obs_qs.filter(urgency__in=['high', 'critical'], created_at__gte=ago30).count()

    neg_emotions = EmotionEntry.objects.filter(
        user=student, emotion__in=['anxious', 'sad', 'angry'],
        created_at__gte=ago14
    ).count()

    last_psych = PsychTestResult.objects.filter(student=student).order_by('-completed_at').first()

    # ── 4. Повышение риска ────────────────────────────────────
    if req.category == 'bullying':
        risk_idx = max(risk_idx, _RISK_ORDER.index('high'))
    if past_30d >= 2:
        risk_idx = min(risk_idx + 1, 3)
    if obs_critical >= 1:
        risk_idx = min(risk_idx + 1, 3)
    if neg_emotions >= 5:
        risk_idx = min(risk_idx + 1, 3)
    if last_psych and last_psych.risk_level == 'high':
        risk_idx = max(risk_idx, _RISK_ORDER.index('medium'))

    final_risk = _RISK_ORDER[risk_idx]

    # ── 5. Теги ───────────────────────────────────────────────
    cat_tag_map = {
        'anxiety': 'тревожность', 'bullying': 'буллинг', 'family': 'семья',
        'study': 'учёба', 'behavior': 'поведение', 'relations': 'отношения',
    }
    tone_tag_map = {
        'fear': 'угроза безопасности', 'aggression': 'агрессия',
        'depression': 'подавленность',  'anxiety': 'тревога',
        'confusion': 'растерянность',
    }
    tags = []
    if cat_tag_map.get(req.category):
        tags.append(cat_tag_map[req.category])
    if tone_tag_map.get(detected_tone):
        tags.append(tone_tag_map[detected_tone])
    seen = set(tags)
    for kw in matched_kw:
        if kw not in seen and len(tags) < 7:
            tags.append(kw)
            seen.add(kw)

    # ── 6. Рекомендации ───────────────────────────────────────
    recs = []
    if detected_tone == 'fear' or req.category == 'bullying':
        recs.append('Провести конфиденциальную беседу о безопасности ученика в коллективе')
        recs.append('Уточнить у классного руководителя наличие видимых конфликтов')
    if detected_tone == 'depression':
        recs.append('При очной встрече оценить глубину состояния, исключить суицидальные мысли')
        recs.append('Рассмотреть регулярные поддерживающие сеансы (2–3 раза в месяц)')
    if detected_tone == 'anxiety':
        recs.append('Выяснить конкретные источники тревоги: учёба, дом, социальная среда')
        recs.append('Предложить техники саморегуляции и работу с мышлением')
    if detected_tone == 'aggression':
        recs.append('Исследовать, не является ли ученик сам объектом давления или буллинга')
        recs.append('Проверить семейный контекст — возможна хроническая стрессовая среда')
    if past_30d >= 2:
        recs.append(f'Повторное обращение за месяц ({past_30d} раза) — взять под регулярный мониторинг')
    if obs_30d > 0:
        recs.append(f'Учитель отправил {obs_30d} сигнал(ов) за 30 дней — запросить детали у педагога')
    if not recs:
        recs.append('Провести вводную беседу, уточнить детали ситуации')
        recs.append('Оценить эмоциональное состояние при личной встрече')

    # ── 7. Сборка текста резюме ───────────────────────────────
    risk_label = {'low': 'Низкий', 'medium': 'Средний',
                  'high': 'Высокий', 'critical': 'Критический'}
    tone_label = {
        'fear': 'Страх / угроза безопасности', 'aggression': 'Агрессия / раздражение',
        'depression': 'Подавленность / депрессивные признаки', 'anxiety': 'Тревожность',
        'confusion': 'Растерянность / дезориентация', 'neutral': 'Нейтральная',
    }
    lines = [
        f'ТОНАЛЬНОСТЬ: {tone_label[detected_tone]}',
        f'УРОВЕНЬ РИСКА: {risk_label[final_risk]}',
        '',
    ]
    history = []
    if past_total:
        history.append(f'Всего предыдущих обращений: {past_total} (за 30 дней: {past_30d})')
    if obs_total:
        history.append(f'Сигналов учителей: {obs_total} (за 30 дней: {obs_30d}, высокой срочности: {obs_critical})')
    if neg_emotions:
        history.append(f'Негативных эмоций в дневнике за 14 дней: {neg_emotions}')
    if last_psych:
        lines_risk = risk_label.get(last_psych.risk_level, last_psych.risk_level)
        history.append(f'Последний тест: группа риска — {lines_risk}')
    if history:
        lines.append('ИСТОРИЯ УЧЕНИКА:')
        lines += [f'• {h}' for h in history]
        lines.append('')

    lines.append('РЕКОМЕНДАЦИИ:')
    lines += [f'{i}. {r}' for i, r in enumerate(recs, 1)]

    # ── 8. Сохраняем ─────────────────────────────────────────
    req.ai_tone        = detected_tone
    req.ai_risk_level  = final_risk
    req.ai_summary     = '\n'.join(lines)
    req.ai_tags        = tags
    req.ai_analyzed_at = now
    req.status         = 'ai_analyzed'
    req.save(update_fields=[
        'ai_tone', 'ai_risk_level', 'ai_summary',
        'ai_tags', 'ai_analyzed_at', 'status',
    ])
    return req


def create_notification(user, notif_type, text, link=''):
    """Создаёт уведомление для пользователя"""
    Notification.objects.create(
        user=user, notif_type=notif_type, text=text, link=link
    )

def get_lang(r): return r.session.get('lang', 'ru')
def set_language(request, lang):
    request.session['lang'] = lang
    return redirect(request.META.get('HTTP_REFERER', '/'))
def get_role(user):
    try: return user.profile.role
    except: return 'student'
def role_redirect(user):
    r = get_role(user)
    if r == 'psychologist': return redirect('psychologist_dashboard')
    if r == 'parent': return redirect('parent_dashboard')
    if r == 'teacher': return redirect('teacher_dashboard')
    return redirect('student_dashboard')

def landing(request):
    return render(request, 'core/home.html', {'lang': get_lang(request)})
def home(request):
    if request.user.is_authenticated: return role_redirect(request.user)
    return landing(request)

def login_view(request):
    lang = get_lang(request); error = None
    if request.user.is_authenticated: return role_redirect(request.user)
    if request.method == 'POST':
        u = authenticate(request, username=request.POST.get('username','').strip(), password=request.POST.get('password',''))
        if u:
            login(request, u); UserProfile.objects.get_or_create(user=u); return role_redirect(u)
        error = 'Неверный логин или пароль / Қате логин немесе пароль'
    return render(request, 'core/login.html', {'lang': lang, 'error': error})

def register_view(request):
    lang = get_lang(request); error = None
    if request.user.is_authenticated: return role_redirect(request.user)
    if request.method == 'POST':
        uname = request.POST.get('username','').strip()
        pwd   = request.POST.get('password','')
        role  = request.POST.get('role','student')
        if not uname or not pwd: error = 'Заполните все поля'
        elif User.objects.filter(username=uname).exists(): error = 'Логин занят / Логин бос емес'
        else:
            u = User.objects.create_user(username=uname, password=pwd,
                first_name=request.POST.get('first_name','').strip(),
                last_name=request.POST.get('last_name','').strip())
            UserProfile.objects.create(user=u, role=role)
            login(request, u); return role_redirect(u)
    return render(request, 'core/register.html', {'lang': lang, 'error': error})

def logout_view(request):
    logout(request); return redirect('landing')

# ── STUDENT ──────────────────────────────────────────────────
@login_required
def student_dashboard(request):
    lang = get_lang(request)
    emotions = EmotionEntry.objects.filter(user=request.user)[:7]
    recent_results = TestResult.objects.filter(user=request.user)[:3]
    # Доступные психологические тесты
    from .models import PsychTest, PsychTestResult
    student_class = getattr(request.user.profile, 'school_class', None)
    available_psych_tests = PsychTest.objects.filter(
        status='active', target_class=student_class
    ) if student_class else PsychTest.objects.none()
    completed_psych_ids = list(PsychTestResult.objects.filter(
        student=request.user
    ).values_list('test_id', flat=True))
    new_psych_tests = [t for t in available_psych_tests if t.id not in completed_psych_ids]

    # Diary streak: consecutive days with at least one emotion entry
    today_date = timezone.now().date()
    streak = 0
    check_day = today_date
    for _ in range(366):
        if EmotionEntry.objects.filter(user=request.user, created_at__date=check_day).exists():
            streak += 1
            check_day -= timedelta(days=1)
        else:
            break

    return render(request, 'core/student/dashboard.html', {
        'lang': lang,
        'new_psych_tests': new_psych_tests, 'emotions': emotions, 'recent_results': recent_results,
        'emotion_count': EmotionEntry.objects.filter(user=request.user).count(),
        'test_count': TestResult.objects.filter(user=request.user).count(),
        'chat_count': ChatMessage.objects.filter(user=request.user).count(),
        'diary_streak': streak,
    })

@login_required
def emotion_diary(request):
    lang = get_lang(request)
    if request.method == 'POST':
        em = request.POST.get('emotion','').strip()
        if em: EmotionEntry.objects.create(user=request.user, emotion=em, event_type=request.POST.get('event_type',''), note=request.POST.get('note','').strip())
        return redirect('emotion_diary')
    entries = EmotionEntry.objects.filter(user=request.user)[:30]
    today = timezone.now().date()
    labels, h, c, a, s = [], [], [], [], []
    for i in range(13,-1,-1):
        day = today - timedelta(days=i)
        labels.append(day.strftime('%d.%m'))
        day_qs = EmotionEntry.objects.filter(user=request.user, created_at__date=day)
        ct = Counter(e.emotion for e in day_qs)
        h.append(ct.get('happy',0)); c.append(ct.get('calm',0))
        a.append(ct.get('anxious',0)); s.append(ct.get('sad',0))
    return render(request, 'core/student/emotion_diary.html', {
        'lang': lang, 'entries': entries,
        'chart_labels': json.dumps(labels),
        'chart_happy': json.dumps(h), 'chart_calm': json.dumps(c),
        'chart_anxious': json.dumps(a), 'chart_sad': json.dumps(s),
    })

@login_required
def test_center(request):
    lang = get_lang(request)
    cats = [
        {'key':'anxiety','name_ru':'Тревожность','name_kz':'Мазасыздық','icon':'brain','desc_ru':'Оцени уровень тревоги','desc_kz':'Мазасыздық деңгейін бағала'},
        {'key':'stress','name_ru':'Стресс','name_kz':'Стресс','icon':'zap','desc_ru':'Как справляешься со стрессом?','desc_kz':'Стресске қалай төтеп бересің?'},
        {'key':'motivation','name_ru':'Мотивация','name_kz':'Мотивация','icon':'rocket','desc_ru':'Уровень мотивации к учёбе','desc_kz':'Оқуға деген мотивация деңгейің'},
        {'key':'social','name_ru':'Социальный','name_kz':'Әлеуметтік','icon':'users','desc_ru':'Комфортно ли в коллективе?','desc_kz':'Ұжымда ыңғайлы ма?'},
    ]
    # Psych-assigned tests for this student's class
    student_class = getattr(request.user.profile, 'school_class', None)
    assigned_tests = PsychTest.objects.filter(
        status='active', target_class=student_class
    ) if student_class else PsychTest.objects.none()
    completed_ids = list(
        PsychTestResult.objects.filter(student=request.user).values_list('test_id', flat=True)
    )
    return render(request, 'core/student/test_center.html', {
        'lang': lang, 'categories': cats,
        'recent_results': TestResult.objects.filter(user=request.user)[:5],
        'assigned_tests': assigned_tests,
        'completed_ids': completed_ids,
    })

QS = {
    'anxiety': [
        {'ru':'Я часто беспокоюсь без причины','kz':'Жиі себепсіз алаңдаймын'},
        {'ru':'Мне трудно расслабиться','kz':'Демалу қиын'},
        {'ru':'Учащённое сердцебиение от волнения','kz':'Толқудан жүрек жиі соғады'},
        {'ru':'Избегаю сложных ситуаций','kz':'Күрделі жағдайлардан аулақ жүремін'},
        {'ru':'Сложно сосредоточиться из-за тревоги','kz':'Алаңдаушылықтан шоғырлану қиын'},
    ],
    'stress': [
        {'ru':'Чувствую себя перегруженным задачами','kz':'Тапсырмалармен шамадан тыс жүктелгенмін'},
        {'ru':'Не хватает времени на отдых','kz':'Демалуға уақытым жетпейді'},
        {'ru':'Раздражаюсь из-за мелочей','kz':'Ұсақ-түйектерге ашуланамын'},
        {'ru':'Головные боли от напряжения','kz':'Кернеуден бас ауырады'},
        {'ru':'Плохо сплю из-за переживаний','kz':'Алаңдаушылықтан нашар ұйықтаймын'},
    ],
    'motivation': [
        {'ru':'Мне интересно учиться','kz':'Оқуды ұнатамын'},
        {'ru':'Ставлю цели и достигаю их','kz':'Мақсат қоямын және оған жетемін'},
        {'ru':'Неудачи не останавливают меня','kz':'Сәтсіздіктер тоқтатпайды'},
        {'ru':'Верю в свои силы','kz':'Өз күшіме сенемін'},
        {'ru':'Нравится узнавать новое','kz':'Жаңа нәрсе үйрену ұнайды'},
    ],
    'social': [
        {'ru':'Легко нахожу общий язык с людьми','kz':'Адамдармен тіл табысамын'},
        {'ru':'Комфортно работать в группе','kz':'Топта жұмыс жасауға жайлы'},
        {'ru':'Могу попросить о помощи','kz':'Көмек сұрай аламын'},
        {'ru':'Есть близкие друзья','kz':'Жақын достарым бар'},
        {'ru':'Умею слушать других','kz':'Басқаларды тыңдай аламын'},
    ],
}

@login_required
def take_test(request, category):
    lang = get_lang(request)
    if request.method == 'POST':
        score = sum(int(request.POST.get(f'q{i}',0)) for i in range(5))
        mx = 20
        if score<=5: ru,kz='Отличный результат! Всё в норме.','Тамаша нәтиже! Бәрі қалыпты.'
        elif score<=10: ru,kz='Умеренный уровень. Стоит обратить внимание.','Орташа деңгей. Назар аудару керек.'
        elif score<=15: ru,kz='Повышенный уровень. Рекомендуем психолога.','Жоғары деңгей. Психолог ұсынылады.'
        else: ru,kz='Высокий уровень. Обратитесь к специалисту.','Өте жоғары. Маманға барыңыз.'
        TestResult.objects.create(user=request.user, category=category, score=score, max_score=mx,
                                  interpretation=ru if lang=='ru' else kz)
        # Если балл высокий — предлагаем запись к психологу
        suggest_appointment = score >= 12
        return render(request, 'core/student/test_result.html',
                      {'lang':lang,'score':score,'max_score':mx,'interp_ru':ru,'interp_kz':kz,
                       'category':category,'suggest_appointment':suggest_appointment})
    return render(request, 'core/student/take_test.html',
                  {'lang':lang,'questions':QS.get(category,QS['anxiety']),'category':category})

@login_required
def ai_chat_view(request):
    return render(request, 'core/student/ai_chat.html', {'lang': get_lang(request)})

@require_POST
def ai_chat_api(request):
    try: data = json.loads(request.body); msg = data.get('message','')
    except: data={}; msg=''
    lang = request.session.get('lang','ru')

    # Ключевые слова для определения тревожного состояния
    concern_keywords_ru = ['стресс','тревог','страшно','боюсь','плохо','не могу','помоги','устал','грустно','плачу','злюсь','не хочу','тяжело','паника','депресси']
    concern_keywords_kz = ['стресс','алаңда','қорқам','жаман','мүмкін емес','көмек','шаршадым','қайғылы','жылаймын','ашулы','қиын','паника']

    msg_lower = msg.lower()
    is_concerning = any(kw in msg_lower for kw in (concern_keywords_kz if lang=='kz' else concern_keywords_ru))

    if is_concerning:
        resp = ('Мен сенің жағдайыңды сезінемін. Психологпен сөйлессең жақсы болар еді. Жазылғың келе ме?' if lang=='kz'
                else 'Я слышу тебя — звучит непросто. Думаю, разговор с психологом поможет. Хочешь записаться?')
        suggest = True
    else:
        pool_ru = ["Я понимаю тебя. Расскажи подробнее — что тебя беспокоит?",
                   "Это непросто. Как давно ты так себя чувствуешь?",
                   "Я здесь, чтобы выслушать. Ты не один/одна в этом.",
                   "Попробуй пройти тест на тревожность — это поможет разобраться.",
                   "Расскажи мне больше — что происходит?"]
        pool_kz = ["Мен сені түсінемін. Толығырақ айт — не алаңдатады?",
                   "Бұл оңай емес. Бұл сезімді қашаннан сезінесің?",
                   "Мен сені тыңдаймын. Сен жалғыз емессің.",
                   "Мазасыздық тестін өту — жағдайды түсінуге көмектеседі.",
                   "Маған көбірек айт — не болып жатыр?"]
        resp = random.choice(pool_kz if lang=='kz' else pool_ru)
        suggest = False

    if request.user.is_authenticated:
        ChatMessage.objects.create(user=request.user, message=msg, is_bot=False)
        ChatMessage.objects.create(user=request.user, message=resp, is_bot=True)

    return JsonResponse({'response': resp, 'suggest_appointment': suggest})

# ── PARENT ───────────────────────────────────────────────────
@login_required
def parent_dashboard(request):
    lang = get_lang(request)
    if get_role(request.user) != 'parent': return role_redirect(request.user)

    from .models import ParentStudent
    import json as json_mod

    def build_children_data(children):
        data = []
        today = timezone.now().date()
        for child in children:
            last_emotions = EmotionEntry.objects.filter(user=child).order_by('-created_at')[:5]
            last_test     = TestResult.objects.filter(user=child).order_by('-created_at').first()
            last_emotion  = last_emotions.first()

            # Sparkline — настроение за 7 дней (1=плохо, 5=хорошо)
            score_map = {'happy':5,'calm':4,'tired':3,'sad':2,'anxious':1,'angry':1}
            sparkline = []
            for i in range(6, -1, -1):
                day = today - timedelta(days=i)
                e = EmotionEntry.objects.filter(
                    user=child, created_at__date=day
                ).order_by('-created_at').first()
                sparkline.append(score_map.get(e.emotion, 3) if e else None)

            data.append({
                'user': child,
                'last_emotion': last_emotion,
                'last_emotions': last_emotions,
                'last_test': last_test,
                'sparkline': json_mod.dumps(sparkline),
                'has_alert': last_emotion and last_emotion.emotion in ['anxious','sad','angry'],
            })
        return data

    children_links = ParentStudent.objects.filter(parent=request.user).select_related('student')
    children = [link.student for link in children_links]
    children_data = build_children_data(children)
    articles = Article.objects.filter(audience='parent')[:6]

    # Заключения психолога по детям родителя
    conclusions = StudentRequest.objects.filter(
        student__in=children,
        is_approved=True,
    ).select_related('student', 'student__profile__school_class', 'assigned_to').order_by('-approved_at')[:10]

    # Непрочитанные уведомления о заключениях
    unread_conclusions = Notification.objects.filter(
        user=request.user,
        notif_type='parent_conclusion',
        is_read=False,
    ).count()

    link_error = link_success = None

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add_child':
            child_username = request.POST.get('child_username', '').strip()
            try:
                child_user = User.objects.get(username=child_username)
                if get_role(child_user) != 'student':
                    link_error = 'Этот пользователь не является учеником / Бұл пайдаланушы оқушы емес'
                elif ParentStudent.objects.filter(parent=request.user, student=child_user).exists():
                    link_error = 'Этот ученик уже привязан / Бұл оқушы бұрыннан қосылған'
                else:
                    ParentStudent.objects.create(parent=request.user, student=child_user)
                    link_success = f'Ученик {child_user.get_full_name() or child_user.username} успешно добавлен!'
                    children_links = ParentStudent.objects.filter(parent=request.user).select_related('student')
                    children = [link.student for link in children_links]
                    children_data = build_children_data(children)
            except User.DoesNotExist:
                link_error = 'Пользователь не найден / Пайдаланушы табылмады'
        elif action == 'remove_child':
            child_id = request.POST.get('child_id')
            ParentStudent.objects.filter(parent=request.user, student_id=child_id).delete()
            return redirect('parent_dashboard')

    return render(request, 'core/parent/dashboard.html', {
        'lang': lang,
        'children_data': children_data,
        'articles': articles,
        'link_error': link_error,
        'link_success': link_success,
        'has_children': bool(children_data),
        'conclusions': conclusions,
        'unread_conclusions': unread_conclusions,
        'notifications_count': Notification.objects.filter(user=request.user, is_read=False).count(),
    })

def articles_view(request):
    lang = get_lang(request)
    audience = request.GET.get('audience', 'parent')

    audience_tabs = [
        ('parent',  'Родителям' if lang != 'kz' else 'Ата-аналарға',  'users'),
        ('student', 'Ученикам'  if lang != 'kz' else 'Оқушыларға',    'backpack'),
        ('teacher', 'Учителям'  if lang != 'kz' else 'Мұғалімдерге',  'graduation-cap'),
    ]

    static_articles = {
        'parent': [
            {'icon':'brain','bg':'linear-gradient(135deg,#FEF3C7,#FDE68A)','title':'Как распознать стресс у ребёнка?','time':5},
            {'icon':'message-circle','bg':'linear-gradient(135deg,#EDE9FE,#DDD6FE)','title':'Как говорить с ребёнком о чувствах','time':7},
            {'icon':'shield','bg':'linear-gradient(135deg,#FEE2E2,#FECACA)','title':'Буллинг — что делать родителям?','time':10},
            {'icon':'book-open','bg':'linear-gradient(135deg,#DBEAFE,#BFDBFE)','title':'Как повысить мотивацию к учёбе?','time':6},
            {'icon':'leaf','bg':'linear-gradient(135deg,#D1FAE5,#A7F3D0)','title':'Техники дыхания — как научить ребёнка','time':4},
            {'icon':'heart','bg':'linear-gradient(135deg,#FDF3E8,#FDE3BC)','title':'Как выстроить доверительные отношения','time':8},
        ],
        'student': [
            {'icon':'wind','bg':'linear-gradient(135deg,#EDE9FE,#DDD6FE)','title':'Как справляться с тревогой','time':5},
            {'icon':'target','bg':'linear-gradient(135deg,#DBEAFE,#BFDBFE)','title':'Тайм-менеджмент для школьника','time':6},
            {'icon':'star','bg':'linear-gradient(135deg,#D1FAE5,#A7F3D0)','title':'Как поверить в себя','time':4},
        ],
        'teacher': [
            {'icon':'eye','bg':'linear-gradient(135deg,#FEF3C7,#FDE68A)','title':'Как заметить тревогу у ученика','time':7},
            {'icon':'shield-check','bg':'linear-gradient(135deg,#EDE9FE,#DDD6FE)','title':'Создание безопасной среды в классе','time':9},
            {'icon':'bar-chart-2','bg':'linear-gradient(135deg,#DBEAFE,#BFDBFE)','title':'Работа с групповой динамикой','time':8},
        ],
    }

    return render(request, 'core/parent/articles.html', {
        'lang': lang,
        'articles': Article.objects.filter(audience=audience),
        'audience': audience,
        'audience_tabs': audience_tabs,
        'static_articles': static_articles.get(audience, []),
    })

# ── PSYCHOLOGIST ──────────────────────────────────────────────
@login_required
def psychologist_dashboard(request):
    lang = get_lang(request)
    if get_role(request.user) != 'psychologist': return role_redirect(request.user)

    from django.db.models import Q
    from .models import SchoolClass

    search       = request.GET.get('search', '').strip()
    class_filter = request.GET.get('class', '').strip()
    active_panel = request.GET.get('panel', 'overview')

    all_classes = sorted(SchoolClass.objects.all(), key=lambda c: _class_sort_key(c.name))

    # Базовый queryset учеников
    students_qs = User.objects.filter(profile__role='student').select_related('profile__school_class')

    if class_filter:
        students_qs = students_qs.filter(profile__school_class__name=class_filter)
    if search:
        students_qs = students_qs.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(username__icontains=search)
        )

    students = sorted(students_qs, key=lambda u: (
        _class_sort_key(u.profile.school_class.name if u.profile.school_class else 'ЯЯ'),
        u.last_name, u.first_name
    ))

    student_ids = [u.id for u in students]

    # Динамическая статистика по выбранным ученикам
    today = timezone.now().date()
    em_keys = ['happy','calm','anxious','sad','angry','tired']
    em_stats = {k: EmotionEntry.objects.filter(
        emotion=k,
        user__id__in=student_ids,
        created_at__date__gte=today - timedelta(days=7)
    ).count() for k in em_keys}

    filtered_emotion_count = EmotionEntry.objects.filter(user__id__in=student_ids).count()
    filtered_test_count    = TestResult.objects.filter(user__id__in=student_ids).count()

    anon_requests = AnonymousRequest.objects.all().order_by('-created_at')[:30]
    # Для regroup нужна сортировка по user, потом по дате
    all_emotions  = EmotionEntry.objects.filter(
        user__id__in=student_ids
    ).select_related('user').order_by('user__id', '-created_at')[:100]

    # ── emotion_stats как объект для шаблона ──
    from types import SimpleNamespace
    emotion_stats_obj = SimpleNamespace(**em_stats)
    total_emotion_week = sum(em_stats.values())

    # ── Группа риска + телефон родителя для каждого ученика ───
    risk_filter = request.GET.get('risk', '').strip()
    students_with_risk = []
    for s in students:
        # Считаем тревожные эмоции за последние 7 дней
        bad_emotions = EmotionEntry.objects.filter(
            user=s,
            emotion__in=['anxious','sad','angry'],
            created_at__date__gte=today - timedelta(days=7)
        ).count()
        total_emotions = EmotionEntry.objects.filter(
            user=s,
            created_at__date__gte=today - timedelta(days=7)
        ).count()

        if bad_emotions >= 3 or (total_emotions > 0 and bad_emotions / total_emotions >= 0.6):
            risk = 'high'
        elif bad_emotions >= 1:
            risk = 'medium'
        else:
            risk = 'low'

        # Телефон родителя
        parent_link = s.parent_links.select_related('parent__profile').first()
        parent_phone = parent_link.parent.profile.phone if parent_link else ''

        students_with_risk.append({
            'user': s,
            'risk': risk,
            'parent_phone': parent_phone,
        })

    # Фильтр по риску
    if risk_filter:
        students_with_risk = [x for x in students_with_risk if x['risk'] == risk_filter]

    return render(request, 'core/psychologist/dashboard.html', {
        'lang': lang,
        'students': students,
        'students_with_risk': students_with_risk,
        'anon_requests': anon_requests,
        'all_emotions': all_emotions,
        'emotion_stats': json.dumps(em_stats),
        'student_count': len(students_with_risk) if risk_filter else len(students),
        'class_count': SchoolClass.objects.count(),
        'request_count': AnonymousRequest.objects.count(),
        'new_request_count': AnonymousRequest.objects.filter(status='new').count(),
        'emotion_count': filtered_emotion_count,
        'test_count': filtered_test_count,
        'active_panel': active_panel,
        'pending_count': Appointment.objects.filter(slot__psychologist=request.user, status='pending').count(),
        'all_classes': all_classes,
        'search': search,
        'class_filter': class_filter,
        'risk_filter': risk_filter,
        'emotion_stats_obj': emotion_stats_obj,
        'total_emotion_week': total_emotion_week,
        'total_students': User.objects.filter(profile__role='student').count(),
        'next_confirmed_apt': Appointment.objects.filter(
            slot__psychologist=request.user,
            status='confirmed',
            slot__date__gte=tz.now().date()
        ).select_related('slot').order_by('slot__date', 'slot__time_start').first(),
    })


def _class_sort_key(name):
    import re
    if not name:
        return (99, '')
    m = re.match(r'^(\d+)\s*([А-ЯA-Z]*)$', name.strip().upper())
    if m:
        return (int(m.group(1)), m.group(2))
    return (99, name)


@login_required
def respond_to_request(request, req_id):
    if get_role(request.user) != 'psychologist': return redirect('landing')
    anon = get_object_or_404(AnonymousRequest, id=req_id)
    if request.method == 'POST':
        resp = request.POST.get('response','').strip()
        if resp: anon.response=resp; anon.status='closed'; anon.save()
    return redirect('psychologist_dashboard')
    """Умная сортировка: '1А'→(1,'А'), '11Б'→(11,'Б'), '9А'→(9,'А')"""
    import re
    if not name:
        return (99, '')
    m = re.match(r'^(\d+)\s*([А-ЯA-Z]*)$', name.strip().upper())
    if m:
        return (int(m.group(1)), m.group(2))
    return (99, name)


# ── TEACHER ──────────────────────────────────────────────────
@login_required
def teacher_dashboard(request):
    lang = get_lang(request)
    if get_role(request.user) != 'teacher': return role_redirect(request.user)

    from django.db.models import Q
    from .models import SchoolClass

    teacher_class = getattr(request.user.profile, 'school_class', None)

    # ── Фильтры ─────────────────────────────────────────────────
    search       = request.GET.get('search', '').strip()
    class_filter = request.GET.get('class', '').strip()

    # ── Классы доступные учителю ─────────────────────────────────
    if teacher_class:
        # Учитель видит только свой класс
        available_classes = [teacher_class]
    else:
        available_classes = sorted(
            SchoolClass.objects.all(),
            key=lambda c: _class_sort_key(c.name)
        )

    # ── Ученики ──────────────────────────────────────────────────
    if teacher_class:
        students = User.objects.filter(
            profile__role='student',
            profile__school_class=teacher_class
        ).select_related('profile__school_class')
    else:
        students = User.objects.filter(
            profile__role='student'
        ).select_related('profile__school_class')

    if class_filter:
        students = students.filter(profile__school_class__name=class_filter)

    if search:
        students = students.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(username__icontains=search)
        )

    students = sorted(students, key=lambda u: (
        _class_sort_key(u.profile.school_class.name if u.profile.school_class else 'ЯЯ'),
        u.last_name, u.first_name
    ))

    student_ids = [u.id for u in students]
    all_emotions = EmotionEntry.objects.filter(user__id__in=student_ids).select_related('user').order_by('-created_at')[:50]
    all_results  = TestResult.objects.filter(user__id__in=student_ids).select_related('user').order_by('-created_at')[:30]
    articles = Article.objects.all()[:6]

    today = timezone.now().date()
    em_keys = ['happy','calm','anxious','sad','angry','tired']
    em_stats = {k: EmotionEntry.objects.filter(
        emotion=k, user__id__in=student_ids,
        created_at__date__gte=today - timedelta(days=7)
    ).count() for k in em_keys}

    # Attention flags: top-3 students with most negative emotions in last 7 days
    attention_flags = []
    for student in list(students)[:50]:
        bad_count = EmotionEntry.objects.filter(
            user=student,
            emotion__in=['anxious', 'sad', 'angry'],
            created_at__date__gte=today - timedelta(days=7)
        ).count()
        if bad_count > 0:
            attention_flags.append({'user': student, 'count': bad_count})
    attention_flags.sort(key=lambda x: -x['count'])
    attention_flags = attention_flags[:3]

    # Class climate: % of positive emotions in last 7 days
    total_em = sum(em_stats.values())
    positive_em = em_stats.get('happy', 0) + em_stats.get('calm', 0)
    class_climate = round(positive_em / total_em * 100) if total_em > 0 else 50

    return render(request, 'core/teacher/dashboard.html', {
        'lang': lang,
        'students': students,
        'all_emotions': all_emotions,
        'all_results': all_results,
        'articles': articles,
        'emotion_stats': json.dumps(em_stats),
        'student_count': len(students),
        'class_count': len(available_classes),
        'emotion_count': len(all_emotions),
        'test_count': len(all_results),
        'teacher_class': teacher_class,
        'available_classes': available_classes,
        'search': search,
        'class_filter': class_filter,
        'attention_flags': attention_flags,
        'class_climate': class_climate,
    })

# ── PUBLIC ANON ───────────────────────────────────────────────
def anonymous_support(request):
    lang = get_lang(request)
    if not request.session.session_key: request.session.create()
    if request.method == 'POST':
        msg = request.POST.get('message','').strip()
        if msg:
            AnonymousRequest.objects.create(session_key=request.session.session_key, message=msg, reason=request.POST.get('category',''))
            # Уведомляем всех психологов
            psychologists = User.objects.filter(profile__role='psychologist')
            for psych in psychologists:
                create_notification(
                    psych,
                    'anonymous_new',
                    f'Новый анонимный запрос: "{msg[:60]}{"..." if len(msg)>60 else ""}"',
                    '/psychologist/'
                )
        return render(request, 'core/anon_success.html', {'lang':lang})
    return render(request, 'core/anonymous_support.html', {'lang':lang})

# ── PROFILE ──────────────────────────────────────────────────

def _get_info_panel_ctx(request, role, lang):
    """Returns context variables required by the role-specific info panel."""
    from .models import SchoolClass, ParentStudent
    import json as _json
    ctx = {}
    today = timezone.now().date()

    if role == 'psychologist':
        ctx['student_count'] = User.objects.filter(profile__role='student').count()
        ctx['class_count'] = SchoolClass.objects.count()
        ctx['new_request_count'] = AnonymousRequest.objects.filter(status='new').count()
        ctx['next_confirmed_apt'] = Appointment.objects.filter(
            slot__psychologist=request.user,
            status='confirmed',
            slot__date__gte=today,
        ).select_related('slot', 'student').order_by('slot__date', 'slot__time_start').first()
        ctx['all_classes'] = sorted(SchoolClass.objects.all(), key=lambda c: _class_sort_key(c.name))
        ctx['search'] = ''
        ctx['class_filter'] = ''

    elif role == 'parent':
        children = [lnk.student for lnk in
                    ParentStudent.objects.filter(parent=request.user).select_related('student')]
        score_map = {'happy': 5, 'calm': 4, 'tired': 3, 'sad': 2, 'anxious': 1, 'angry': 1}
        children_data = []
        for child in children:
            last_emotion = EmotionEntry.objects.filter(user=child).order_by('-created_at').first()
            sparkline = []
            for i in range(6, -1, -1):
                day = today - timedelta(days=i)
                e = EmotionEntry.objects.filter(user=child, created_at__date=day).order_by('-created_at').first()
                sparkline.append(score_map.get(e.emotion, 3) if e else None)
            children_data.append({
                'user': child,
                'last_emotion': last_emotion,
                'last_emotions': EmotionEntry.objects.filter(user=child).order_by('-created_at')[:5],
                'last_test': TestResult.objects.filter(user=child).order_by('-created_at').first(),
                'sparkline': _json.dumps(sparkline),
                'has_alert': bool(last_emotion and last_emotion.emotion in ['anxious', 'sad', 'angry']),
            })
        ctx['children_data'] = children_data
        ctx['articles'] = Article.objects.filter(audience='parent')[:6]
        ctx['notifications_count'] = Notification.objects.filter(user=request.user, is_read=False).count()

    elif role == 'student':
        streak = 0
        check_day = today
        for _ in range(366):
            if EmotionEntry.objects.filter(user=request.user, created_at__date=check_day).exists():
                streak += 1
                check_day -= timedelta(days=1)
            else:
                break
        ctx['emotion_count'] = EmotionEntry.objects.filter(user=request.user).count()
        ctx['test_count'] = TestResult.objects.filter(user=request.user).count()
        ctx['chat_count'] = ChatMessage.objects.filter(user=request.user).count()
        ctx['diary_streak'] = streak

    elif role == 'teacher':
        from .models import SchoolClass
        teacher_class = getattr(request.user.profile, 'school_class', None)
        if teacher_class:
            students = list(User.objects.filter(
                profile__role='student', profile__school_class=teacher_class
            ).select_related('profile__school_class')[:50])
            available_classes = [teacher_class]
        else:
            students = list(User.objects.filter(profile__role='student').select_related('profile__school_class')[:50])
            available_classes = sorted(SchoolClass.objects.all(), key=lambda c: _class_sort_key(c.name))

        student_ids = [u.id for u in students]
        em_keys = ['happy', 'calm', 'anxious', 'sad', 'angry', 'tired']
        em_stats = {k: EmotionEntry.objects.filter(
            emotion=k, user__id__in=student_ids,
            created_at__date__gte=today - timedelta(days=7)
        ).count() for k in em_keys}

        attention_flags = []
        for student in students:
            bad = EmotionEntry.objects.filter(
                user=student, emotion__in=['anxious', 'sad', 'angry'],
                created_at__date__gte=today - timedelta(days=7)
            ).count()
            if bad > 0:
                attention_flags.append({'user': student, 'count': bad})
        attention_flags.sort(key=lambda x: -x['count'])

        total_em = sum(em_stats.values())
        positive_em = em_stats.get('happy', 0) + em_stats.get('calm', 0)

        ctx['student_count'] = len(students)
        ctx['emotion_count'] = EmotionEntry.objects.filter(user__id__in=student_ids).count()
        ctx['test_count'] = TestResult.objects.filter(user__id__in=student_ids).count()
        ctx['teacher_class'] = teacher_class
        ctx['available_classes'] = available_classes
        ctx['class_filter'] = ''
        ctx['attention_flags'] = attention_flags[:3]
        ctx['class_climate'] = round(positive_em / total_em * 100) if total_em > 0 else 50

    return ctx


@login_required
def profile_view(request):
    lang = get_lang(request)
    profile = request.user.profile
    role = profile.role
    success = None
    error = None

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'update_info':
            first_name   = request.POST.get('first_name', '').strip()
            last_name    = request.POST.get('last_name', '').strip()
            email        = request.POST.get('email', '').strip()
            phone        = request.POST.get('phone', '').strip()
            new_username = request.POST.get('username', '').strip()

            if not new_username:
                error = 'Логин не может быть пустым'
            elif new_username != request.user.username and User.objects.filter(username=new_username).exists():
                error = 'Этот логин уже занят / Бұл логин бос емес'
            else:
                request.user.first_name = first_name
                request.user.last_name  = last_name
                request.user.email      = email
                request.user.username   = new_username
                request.user.save()
                profile.phone = phone
                profile.save()
                success = 'info'

        elif action == 'change_password':
            old_pwd  = request.POST.get('old_password', '')
            new_pwd  = request.POST.get('new_password', '')
            new_pwd2 = request.POST.get('new_password2', '')

            if not request.user.check_password(old_pwd):
                error = 'Текущий пароль неверный / Ағымдағы құпия сөз қате'
            elif len(new_pwd) < 6:
                error = 'Новый пароль минимум 6 символов / Жаңа құпия сөз кемінде 6 таңба'
            elif new_pwd != new_pwd2:
                error = 'Пароли не совпадают / Құпия сөздер сәйкес емес'
            else:
                request.user.set_password(new_pwd)
                request.user.save()
                from django.contrib.auth import update_session_auth_hash
                update_session_auth_hash(request, request.user)
                success = 'password'

    return render(request, 'core/profile.html', {
        'lang': lang,
        'profile': profile,
        'success': success,
        'error': error,
    })

# ── APPOINTMENTS ─────────────────────────────────────────────
from .models import PsychologistSchedule, Appointment
from django.utils import timezone as tz

@login_required
def book_appointment(request):
    """Страница записи к психологу — автослоты пн-пт 09-18"""
    from datetime import date, timedelta, datetime, time as dtime
    lang = get_lang(request)
    reason = request.GET.get('reason', '')

    today = tz.now().date()

    # Получаем всех психологов
    psychologists = User.objects.filter(profile__role='psychologist')

    # Генерируем свободные слоты на ближайшие 2 недели (пн-пт 09-18)
    free_slots = []
    for offset in range(0, 15):
        day = today + timedelta(days=offset)
        if day.weekday() >= 5:  # Пропускаем выходные
            continue
        for hour in range(9, 18):
            time_start = dtime(hour, 0)
            time_end   = dtime(hour + 1, 0)
            for psych in psychologists:
                # Проверяем не заблокирован ли слот психологом
                is_blocked = PsychologistSchedule.objects.filter(
                    psychologist=psych,
                    date=day,
                    time_start=time_start,
                    is_available=False
                ).exists()
                # Проверяем нет ли уже записи
                is_booked = Appointment.objects.filter(
                    slot__psychologist=psych,
                    slot__date=day,
                    slot__time_start=time_start,
                    status__in=['pending','confirmed']
                ).exists()

                if not is_blocked and not is_booked:
                    free_slots.append({
                        'psychologist': psych,
                        'date': day,
                        'time_start': time_start,
                        'time_end': time_end,
                        'slot_key': f"{psych.id}_{day}_{hour}",
                    })

    # Группируем по дате
    from itertools import groupby
    from operator import itemgetter
    slots_by_date = {}
    for slot in free_slots:
        d = slot['date']
        if d not in slots_by_date:
            slots_by_date[d] = []
        slots_by_date[d].append(slot)

    if request.method == 'POST':
        # slot_key: psych_id_date_hour
        slot_key = request.POST.get('slot_key', '')
        note = request.POST.get('note', '').strip()
        reason = request.POST.get('reason', '')  # берём из формы, не из GET
        try:
            parts = slot_key.split('_')
            psych_id = int(parts[0])
            slot_date = date.fromisoformat(parts[1])
            slot_hour = int(parts[2])
            time_start = dtime(slot_hour, 0)
            time_end   = dtime(slot_hour + 1, 0)
            psych = User.objects.get(id=psych_id, profile__role='psychologist')

            # Создаём или получаем слот в PsychologistSchedule
            slot_obj, _ = PsychologistSchedule.objects.get_or_create(
                psychologist=psych,
                date=slot_date,
                time_start=time_start,
                defaults={'time_end': time_end, 'is_available': True}
            )

            # Создаём запись
            Appointment.objects.create(
                student=request.user,
                slot=slot_obj,
                reason=reason,
                student_note=note,
            )
            # Уведомляем психолога
            create_notification(
                psych,
                'appointment_new',
                f'Новая заявка на приём от {request.user.get_full_name() or request.user.username} на {slot_date.strftime("%d.%m.%Y")} в {time_start.strftime("%H:%M")}',
                '/psychologist/appointments/'
            )
            return redirect('my_appointments')
        except Exception as e:
            pass

    return render(request, 'core/student/book_appointment.html', {
        'lang': lang,
        'slots_by_date': slots_by_date,
        'reason': reason,
    })


@login_required
def my_appointments(request):
    """Мои записи — для ученика"""
    lang = get_lang(request)
    appointments = Appointment.objects.filter(
        student=request.user
    ).select_related('slot', 'slot__psychologist').order_by('-created_at')

    if request.method == 'POST' and request.POST.get('action') == 'cancel':
        apt_id = request.POST.get('appointment_id')
        try:
            apt = Appointment.objects.get(id=apt_id, student=request.user, status='pending')
            apt.status = 'cancelled'
            apt.save()
            # Освобождаем слот
            apt.slot.is_available = True
            apt.slot.save()
        except Appointment.DoesNotExist:
            pass
        return redirect('my_appointments')

    return render(request, 'core/student/my_appointments.html', {
        'lang': lang,
        'appointments': appointments,
    })


@login_required
def psychologist_appointments(request):
    """Управление записями — для психолога"""
    lang = get_lang(request)
    if get_role(request.user) != 'psychologist':
        return role_redirect(request.user)

    today = tz.now().date()

    # Входящие заявки (ожидают)
    pending = Appointment.objects.filter(
        slot__psychologist=request.user,
        status='pending'
    ).select_related('student', 'slot').order_by('slot__date', 'slot__time_start')

    # Подтверждённые предстоящие (дата >= сегодня)
    confirmed = Appointment.objects.filter(
        slot__psychologist=request.user,
        status='confirmed',
        slot__date__gte=today
    ).select_related('student', 'slot').order_by('slot__date', 'slot__time_start')

    # Архив — все прошедшие (подтверждённые прошлые + отклонённые + отменённые)
    archive = Appointment.objects.filter(
        slot__psychologist=request.user,
    ).filter(
        # Прошедшие подтверждённые ИЛИ отклонённые/отменённые
        **{}
    ).exclude(
        status='pending'
    ).exclude(
        status='confirmed', slot__date__gte=today
    ).select_related('student', 'slot').order_by('-slot__date', '-slot__time_start')[:50]

    if request.method == 'POST':
        apt_id = request.POST.get('appointment_id')
        action = request.POST.get('action')
        note   = request.POST.get('note', '').strip()
        try:
            apt = Appointment.objects.get(id=apt_id, slot__psychologist=request.user)
            if action == 'confirm':
                apt.status = 'confirmed'
                apt.slot.is_available = False
                apt.slot.save()
                # Уведомление ученику
                create_notification(
                    apt.student,
                    'appointment_confirmed',
                    f'Ваша запись к психологу {request.user.get_full_name()} на {apt.slot.date.strftime("%d.%m.%Y")} в {apt.slot.time_start.strftime("%H:%M")} подтверждена!',
                    '/my-appointments/'
                )
            elif action == 'reject':
                apt.status = 'rejected'
                apt.slot.is_available = True
                apt.slot.save()
                # Уведомление ученику об отклонении
                create_notification(
                    apt.student,
                    'appointment_rejected',
                    f'К сожалению, ваша запись к психологу на {apt.slot.date.strftime("%d.%m.%Y")} в {apt.slot.time_start.strftime("%H:%M")} была отклонена. {("Причина: " + note) if note else "Попробуйте выбрать другое время."}',
                    '/my-appointments/'
                )
            apt.psychologist_note = note
            apt.save()
        except Appointment.DoesNotExist:
            pass
        return redirect('psychologist_appointments')

    from .models import SchoolClass, UserProfile, AnonymousRequest
    return render(request, 'core/psychologist/appointments.html', {
        'lang': lang,
        'pending': pending,
        'confirmed': confirmed,
        'archive': archive,
        'today': today,
        'student_count': UserProfile.objects.filter(role='student').count(),
        'class_count': SchoolClass.objects.count(),
        'new_request_count': AnonymousRequest.objects.filter(status='new').count(),
        'all_classes': sorted(SchoolClass.objects.all(), key=lambda c: _class_sort_key(c.name)),
    })


@login_required
def manage_schedule(request):
    """Психолог управляет расписанием — пн-пт 09-18 всегда свободны, отмечает занятые"""
    from datetime import date, timedelta, datetime
    lang = get_lang(request)
    if get_role(request.user) != 'psychologist':
        return role_redirect(request.user)

    today = tz.now().date()
    hours = list(range(9, 18))  # 09:00 - 17:00

    # Определяем текущую неделю
    week_offset = int(request.GET.get('week', 0))
    week_start = today - timedelta(days=today.weekday()) + timedelta(weeks=week_offset)
    week_end = week_start + timedelta(days=4)

    # Дни недели (пн-пт)
    week_days_dates = [week_start + timedelta(days=i) for i in range(5)]

    # Занятые слоты психолога (is_available=False — заблокировано им самим)
    blocked = PsychologistSchedule.objects.filter(
        psychologist=request.user,
        is_available=False,
        date__range=[week_start, week_end]
    ).values_list('date', 'time_start')
    blocked_set = set((b[0], b[1].hour) for b in blocked)

    # Подтверждённые записи
    confirmed_apts = Appointment.objects.filter(
        slot__psychologist=request.user,
        status='confirmed',
        slot__date__range=[week_start, week_end]
    ).select_related('slot', 'student')
    confirmed_set = set()
    confirmed_info = {}
    for apt in confirmed_apts:
        key = (apt.slot.date, apt.slot.time_start.hour)
        confirmed_set.add(key)
        confirmed_info[key] = apt.student.get_full_name() or apt.student.username

    # Отклонённые записи (для архива в календаре)
    rejected_apts = Appointment.objects.filter(
        slot__psychologist=request.user,
        status='rejected',
        slot__date__range=[week_start, week_end]
    ).select_related('slot')
    rejected_set = set((apt.slot.date, apt.slot.time_start.hour) for apt in rejected_apts)

    # Строим данные для шаблона — матрица часы x дни
    calendar_rows = []
    for hour in hours:
        row = {'hour': hour, 'cells': []}
        for d in week_days_dates:
            is_blocked = (d, hour) in blocked_set
            is_confirmed = (d, hour) in confirmed_set
            is_rejected = (d, hour) in rejected_set
            student_name = confirmed_info.get((d, hour), '')
            cell = {
                'date': d,
                'hour': hour,
                'date_str': d.isoformat(),
                'is_blocked': is_blocked,
                'is_confirmed': is_confirmed,
                'is_rejected': is_rejected,
                'is_past': d < today,
                'student_name': student_name,
            }
            row['cells'].append(cell)
        calendar_rows.append(row)

    error = None
    success = None

    if request.method == 'POST':
        action = request.POST.get('action')
        date_str = request.POST.get('date')
        hour_str = request.POST.get('hour')

        if date_str and hour_str:
            slot_date = date.fromisoformat(date_str)
            slot_hour = int(hour_str)
            time_start = datetime.strptime(f"{slot_hour}:00", "%H:%M").time()
            time_end   = datetime.strptime(f"{slot_hour+1}:00", "%H:%M").time()

            if action == 'block':
                # Блокируем слот
                PsychologistSchedule.objects.get_or_create(
                    psychologist=request.user,
                    date=slot_date,
                    time_start=time_start,
                    defaults={'time_end': time_end, 'is_available': False}
                )
                obj, _ = PsychologistSchedule.objects.get_or_create(
                    psychologist=request.user,
                    date=slot_date,
                    time_start=time_start,
                    defaults={'time_end': time_end}
                )
                obj.is_available = False
                obj.save()
                success = 'Слот заблокирован'

            elif action == 'free':
                # Освобождаем слот
                PsychologistSchedule.objects.filter(
                    psychologist=request.user,
                    date=slot_date,
                    time_start=time_start,
                    is_available=False
                ).delete()
                success = 'Слот освобождён'

        return redirect(f"{request.path}?week={week_offset}")

    # Заголовки дней
    week_days = [{'date': d, 'is_today': d == today, 'is_past': d < today} for d in week_days_dates]

    from .models import SchoolClass, UserProfile, AnonymousRequest
    return render(request, 'core/psychologist/schedule.html', {
        'lang': lang,
        'week_days': week_days,
        'calendar_rows': calendar_rows,
        'week_start': week_start,
        'week_end': week_end,
        'prev_week': week_offset - 1,
        'next_week': week_offset + 1,
        'current_week': 0,
        'error': error,
        'success': success,
        'today': today,
        'student_count': UserProfile.objects.filter(role='student').count(),
        'class_count': SchoolClass.objects.count(),
        'new_request_count': AnonymousRequest.objects.filter(status='new').count(),
        'all_classes': sorted(SchoolClass.objects.all(), key=lambda c: _class_sort_key(c.name)),
    })


# ── NOTIFICATIONS ─────────────────────────────────────────────
@login_required
def notifications_view(request):
    lang = get_lang(request)
    notifs = Notification.objects.filter(user=request.user)[:30]
    # Отмечаем все как прочитанные
    Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
    return render(request, 'core/notifications.html', {'lang': lang, 'notifications': notifs})


@login_required
def notifications_count(request):
    count = Notification.objects.filter(user=request.user, is_read=False).count()
    return JsonResponse({'count': count})


# ── ADMIN IMPORT ──────────────────────────────────────────────
import tempfile, os
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.http import FileResponse

@staff_member_required
def admin_import_view(request):
    """Страница импорта пользователей из Excel (только для staff/admin)"""
    lang = get_lang(request)
    result = None

    if request.method == 'POST' and request.FILES.get('excel_file'):
        import_file = request.FILES['excel_file']
        dry_run = bool(request.POST.get('dry_run'))

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            for chunk in import_file.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        try:
            result = run_import(tmp_path, dry_run=dry_run)
            # Сохраняем путь к файлу паролей в сессии
            if result.get('output_file'):
                request.session['import_result_path'] = result['output_file']
        finally:
            os.unlink(tmp_path)

    return render(request, 'core/admin_import.html', {'lang': lang, 'result': result})



def run_import(file_path, dry_run=False):
    """Web версия импорта — вызывается из admin_import_view"""
    import subprocess, sys, json, tempfile, os
    from django.contrib.auth.models import User
    from .models import UserProfile, SchoolClass, ParentStudent
    import re, secrets, string
    from openpyxl import load_workbook
    from django.db import transaction

    def next_sid():
        existing = User.objects.filter(username__regex=r'^id\d+$').values_list('username', flat=True)
        nums = [int(u[2:]) for u in existing if u[2:].isdigit()]
        return f'id{(max(nums)+1 if nums else 1):04d}'

    def gen_pwd(base=''): return f'{base}_2026' if base else ''.join(secrets.choice(string.ascii_letters+string.digits) for _ in range(10))

    def mk_uname(name, prefix=''):
        parts = re.sub(r'[^\w\s]','',name.lower()).split()
        base = prefix + '_'.join(parts[:2]) if parts else prefix+'user'
        u = base; c = 1
        while User.objects.filter(username=u).exists(): u=f'{base}_{c}'; c+=1
        return u

    try:
        wb = load_workbook(file_path)
    except Exception as e:
        return {'teachers':0,'students':0,'parents':0,'skipped':0,'errors':[str(e)]}

    ws = wb['Импорт класса'] if 'Импорт класса' in wb.sheetnames else wb.active
    stats = {'teachers':0,'students':0,'parents':0,'skipped':0}
    errors = []; created = []
    teacher_cache = {}; class_cache = {}

    try:
        with transaction.atomic():
            sid = transaction.savepoint()
            for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
                if not any(row): continue
                try:
                    cn=str(row[0] or '').strip(); sn=str(row[1] or '').strip()
                    se=str(row[2] or '').strip() if len(row)>2 else ''
                    tn=str(row[3] or '').strip() if len(row)>3 else ''
                    pn=str(row[4] or '').strip() if len(row)>4 else ''
                    pp_raw=row[5] if len(row)>5 else ''
                    pp=str(int(pp_raw)).strip() if isinstance(pp_raw,(int,float)) else str(pp_raw or '').strip()
                    pe=str(row[6] or '').strip() if len(row)>6 else ''
                    if not cn or not sn: continue

                    if cn not in class_cache:
                        sc,_=SchoolClass.objects.get_or_create(name=cn); class_cache[cn]=sc
                    school_class=class_cache[cn]

                    # Teacher
                    if tn and tn not in teacher_cache:
                        parts=tn.split()
                        tq=User.objects.filter(profile__role='teacher',first_name=parts[0] if parts else '')
                        if len(parts)>1: tq=tq.filter(last_name__icontains=parts[1])
                        if tq.exists():
                            teacher=tq.first(); teacher_cache[tn]=teacher
                        else:
                            tu=mk_uname(tn); tp=gen_pwd()
                            teacher=User.objects.create_user(tu,'',tp,first_name=parts[0] if parts else '',last_name=' '.join(parts[1:]) if len(parts)>1 else '')
                            pr,_=UserProfile.objects.get_or_create(user=teacher)
                            pr.role='teacher'; pr.school_class=school_class; pr.save()
                            teacher_cache[tn]=teacher; stats['teachers']+=1
                            created.append({'role':'Учитель','full_name':tn,'username':tu,'password':tp,'class':cn})

                    # Student
                    parts=sn.split()
                    sq=User.objects.filter(profile__role='student',profile__school_class=school_class,first_name=parts[0] if parts else '')
                    if len(parts)>1: sq=sq.filter(last_name__icontains=parts[1])
                    if sq.exists():
                        student=sq.first(); stats['skipped']+=1
                    else:
                        sid_val=next_sid(); sp=gen_pwd(sid_val)
                        student=User.objects.create_user(sid_val,se,sp,first_name=parts[0] if parts else '',last_name=' '.join(parts[1:]) if len(parts)>1 else '')
                        pr,_=UserProfile.objects.get_or_create(user=student)
                        pr.role='student'; pr.school_class=school_class; pr.save()
                        stats['students']+=1
                        created.append({'role':'Ученик','full_name':sn,'username':sid_val,'password':sp,'email':se,'class':cn,'student_id':sid_val})

                    # Parent
                    if pn and pp:
                        pp_clean=re.sub(r'[\s\-\(\)]','',pp)
                        pp_clean=re.sub(r'[\s\-\(\)\+]','',pp)
                        pq=UserProfile.objects.filter(role='parent',phone__icontains=pp_clean[-7:]).select_related('user') if len(pp_clean)>=7 else UserProfile.objects.none()
                        if pq.exists():
                            parent=pq.first().user
                            ParentStudent.objects.get_or_create(parent=parent,student=student)
                        else:
                            pu=mk_uname(pn,'parent_'); pp2=gen_pwd()
                            pparts=pn.split()
                            parent=User.objects.create_user(pu,pe,pp2,first_name=pparts[0] if pparts else '',last_name=' '.join(pparts[1:]) if len(pparts)>1 else '')
                            pr,_=UserProfile.objects.get_or_create(user=parent)
                            pr.role='parent'; pr.phone=pp; pr.save()
                            ParentStudent.objects.get_or_create(parent=parent,student=student)
                            stats['parents']+=1
                            created.append({'role':'Родитель','full_name':pn,'username':pu,'password':pp2,'email':pe,'phone':pp,'children':student.username})
                except Exception as e:
                    errors.append(f'Строка {row_num}: {e}')

            if dry_run: transaction.savepoint_rollback(sid)
    except Exception as e:
        errors.append(f'Критическая ошибка: {e}')
        return {**stats,'errors':errors}

    output_file = _write_passwords_excel(created) if created and not dry_run else None
    return {**stats,'errors':errors,'output_file':output_file,'dry_run':dry_run}


def _write_passwords_excel(users):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import tempfile

    wb = Workbook(); ws = wb.active; ws.title = 'Логины и пароли'
    h_fill = PatternFill('solid', fgColor='4A7C59')
    headers = ['Роль','ФИО','Логин','Пароль','Email','Телефон','Класс','ID ученика']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = h_fill
    colors = {'Учитель':'EBF3EE','Ученик':'EBF4FA','Родитель':'FDF3E8'}
    for r, u in enumerate(users, 2):
        data = [u.get('role'),u.get('full_name'),u.get('username'),u.get('password'),
                u.get('email'),u.get('phone',''),u.get('class',''),u.get('student_id','')]
        fill = PatternFill('solid', fgColor=colors.get(u.get('role',''),'FFFFFF'))
        for c, v in enumerate(data, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = fill
    for c, w in enumerate([12,30,20,20,30,18,10,14], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = 'A2'

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False, prefix='import_result_')
    wb.save(tmp.name)
    return tmp.name


@staff_member_required
def download_import_template(request):
    """Скачать шаблон Excel"""
    template_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'import_template.xlsx')
    if not os.path.exists(template_path):
        from django.http import HttpResponse
        return HttpResponse('Шаблон не найден', status=404)
    return FileResponse(open(template_path, 'rb'),
        as_attachment=True, filename='import_template.xlsx')


@staff_member_required
def download_import_result(request):
    """Скачать результат импорта"""
    result_path = request.session.get('import_result_path')
    if result_path and os.path.exists(result_path):
        return FileResponse(open(result_path, 'rb'),
            as_attachment=True, filename='import_result.xlsx')
    from django.http import HttpResponse
    return HttpResponse('Файл не найден', status=404)


# ── PARENT BOOK APPOINTMENT ────────────────────────────────────
@login_required
def parent_book_appointment(request):
    """Родитель записывает ребёнка к психологу"""
    from datetime import date, timedelta, time as dtime
    lang = get_lang(request)
    if get_role(request.user) != 'parent': return role_redirect(request.user)

    from .models import ParentStudent

    # Дети родителя
    children_links = ParentStudent.objects.filter(parent=request.user).select_related('student')
    children = [link.student for link in children_links]

    selected_child_id = request.GET.get('child', str(children[0].id) if children else '')
    selected_child = None
    for child in children:
        if str(child.id) == selected_child_id:
            selected_child = child
            break

    # Причины обращения
    reason_choices = [
        ('anxiety',  'Тревога / Стресс'),
        ('bullying', 'Буллинг'),
        ('family',   'Семейные проблемы'),
        ('study',    'Проблемы с учёбой'),
        ('behavior', 'Поведение'),
        ('other',    'Другое'),
    ]

    success = error = None
    slots_by_date = {}

    if selected_child:
        # Генерируем слоты
        today = tz.now().date()
        psychologists = User.objects.filter(profile__role='psychologist')
        free_slots = []
        for offset in range(0, 15):
            day = today + timedelta(days=offset)
            if day.weekday() >= 5:
                continue
            for hour in range(9, 18):
                time_start = dtime(hour, 0)
                time_end   = dtime(hour + 1, 0)
                for psych in psychologists:
                    is_blocked = PsychologistSchedule.objects.filter(
                        psychologist=psych, date=day,
                        time_start=time_start, is_available=False
                    ).exists()
                    is_booked = Appointment.objects.filter(
                        slot__psychologist=psych, slot__date=day,
                        slot__time_start=time_start,
                        status__in=['pending','confirmed']
                    ).exists()
                    if not is_blocked and not is_booked:
                        free_slots.append({
                            'psychologist': psych,
                            'date': day,
                            'time_start': time_start,
                            'time_end': time_end,
                            'slot_key': f"{psych.id}_{day}_{hour}",
                        })

        for slot in free_slots:
            d = slot['date']
            if d not in slots_by_date:
                slots_by_date[d] = []
            slots_by_date[d].append(slot)

    if request.method == 'POST':
        child_id  = request.POST.get('child_id')
        slot_key  = request.POST.get('slot_key', '')
        note      = request.POST.get('note', '').strip()
        reason    = request.POST.get('reason', 'other')

        try:
            child_user = User.objects.get(id=child_id)
            # Проверяем что это действительно ребёнок этого родителя
            if not ParentStudent.objects.filter(parent=request.user, student=child_user).exists():
                error = 'Ошибка доступа'
            else:
                parts = slot_key.split('_')
                psych_id   = int(parts[0])
                slot_date  = date.fromisoformat(parts[1])
                slot_hour  = int(parts[2])
                time_start = dtime(slot_hour, 0)
                time_end   = dtime(slot_hour + 1, 0)
                psych = User.objects.get(id=psych_id, profile__role='psychologist')

                slot_obj, _ = PsychologistSchedule.objects.get_or_create(
                    psychologist=psych, date=slot_date, time_start=time_start,
                    defaults={'time_end': time_end, 'is_available': True}
                )
                Appointment.objects.create(
                    student=child_user, slot=slot_obj,
                    reason=reason, student_note=note,
                )
                # Уведомление психологу
                create_notification(
                    psych, 'appointment_new',
                    f'Новая заявка от родителя: {child_user.get_full_name() or child_user.username} на {slot_date.strftime("%d.%m.%Y")} в {time_start.strftime("%H:%M")}',
                    '/psychologist/appointments/'
                )
                # Уведомление ученику
                create_notification(
                    child_user, 'appointment_new',
                    f'Родитель записал тебя к психологу на {slot_date.strftime("%d.%m.%Y")} в {time_start.strftime("%H:%M")}',
                    '/my-appointments/'
                )
                success = f'{"Бала жазылды!" if lang == "kz" else "Ребёнок записан!"} {slot_date.strftime("%d.%m.%Y")} {"сағат" if lang == "kz" else "в"} {time_start.strftime("%H:%M")}'
        except Exception as e:
            error = str(e)

    return render(request, 'core/parent/book_appointment.html', {
        'lang': lang,
        'children': children,
        'selected_child': selected_child,
        'selected_child_id': selected_child_id,
        'slots_by_date': slots_by_date,
        'reason_choices': reason_choices,
        'success': success,
        'error': error,
    })


# ── TEACHER BOOK APPOINTMENT ───────────────────────────────────
@login_required
def teacher_book_appointment(request):
    """Учитель записывает ученика своего класса к психологу"""
    from datetime import date, timedelta, time as dtime
    lang = get_lang(request)
    if get_role(request.user) != 'teacher': return role_redirect(request.user)

    teacher_class = getattr(request.user.profile, 'school_class', None)

    # Ученики учителя: свой класс, либо все (если класс не задан)
    if teacher_class:
        students_qs = User.objects.filter(
            profile__role='student',
            profile__school_class=teacher_class,
        ).select_related('profile__school_class')
    else:
        students_qs = User.objects.filter(
            profile__role='student',
        ).select_related('profile__school_class')
    students = list(students_qs.order_by('last_name', 'first_name'))

    selected_student_id = request.GET.get('child', str(students[0].id) if students else '')
    selected_student = None
    for s in students:
        if str(s.id) == selected_student_id:
            selected_student = s
            break

    reason_choices = [
        ('anxiety',  'Тревога / Стресс'),
        ('bullying', 'Буллинг'),
        ('family',   'Семейные проблемы'),
        ('study',    'Проблемы с учёбой'),
        ('behavior', 'Поведение'),
        ('other',    'Другое'),
    ]

    success = error = None
    slots_by_date = {}

    if selected_student:
        today = tz.now().date()
        psychologists = User.objects.filter(profile__role='psychologist')
        free_slots = []
        for offset in range(0, 15):
            day = today + timedelta(days=offset)
            if day.weekday() >= 5:
                continue
            for hour in range(9, 18):
                time_start = dtime(hour, 0)
                time_end   = dtime(hour + 1, 0)
                for psych in psychologists:
                    is_blocked = PsychologistSchedule.objects.filter(
                        psychologist=psych, date=day,
                        time_start=time_start, is_available=False
                    ).exists()
                    is_booked = Appointment.objects.filter(
                        slot__psychologist=psych, slot__date=day,
                        slot__time_start=time_start,
                        status__in=['pending','confirmed']
                    ).exists()
                    if not is_blocked and not is_booked:
                        free_slots.append({
                            'psychologist': psych,
                            'date': day,
                            'time_start': time_start,
                            'time_end': time_end,
                            'slot_key': f"{psych.id}_{day}_{hour}",
                        })
        for slot in free_slots:
            d = slot['date']
            slots_by_date.setdefault(d, []).append(slot)

    if request.method == 'POST':
        child_id  = request.POST.get('child_id')
        slot_key  = request.POST.get('slot_key', '')
        note      = request.POST.get('note', '').strip()
        reason    = request.POST.get('reason', 'other')

        try:
            student_user = User.objects.get(id=child_id, profile__role='student')
            # Проверка: ученик должен быть в классе учителя (если задан)
            if teacher_class and student_user.profile.school_class_id != teacher_class.id:
                error = 'Ошибка доступа: ученик не из вашего класса'
            else:
                parts = slot_key.split('_')
                psych_id   = int(parts[0])
                slot_date  = date.fromisoformat(parts[1])
                slot_hour  = int(parts[2])
                time_start = dtime(slot_hour, 0)
                time_end   = dtime(slot_hour + 1, 0)
                psych = User.objects.get(id=psych_id, profile__role='psychologist')

                slot_obj, _ = PsychologistSchedule.objects.get_or_create(
                    psychologist=psych, date=slot_date, time_start=time_start,
                    defaults={'time_end': time_end, 'is_available': True}
                )
                Appointment.objects.create(
                    student=student_user, slot=slot_obj,
                    reason=reason, student_note=note,
                )
                teacher_name = request.user.get_full_name() or request.user.username
                create_notification(
                    psych, 'appointment_new',
                    f'Новая заявка от учителя {teacher_name}: {student_user.get_full_name() or student_user.username} на {slot_date.strftime("%d.%m.%Y")} в {time_start.strftime("%H:%M")}',
                    '/psychologist/appointments/'
                )
                create_notification(
                    student_user, 'appointment_new',
                    f'Учитель записал тебя к психологу на {slot_date.strftime("%d.%m.%Y")} в {time_start.strftime("%H:%M")}',
                    '/my-appointments/'
                )
                success = f'{"Оқушы жазылды!" if lang == "kz" else "Ученик записан!"} {slot_date.strftime("%d.%m.%Y")} {"сағат" if lang == "kz" else "в"} {time_start.strftime("%H:%M")}'
        except Exception as e:
            error = str(e)

    return render(request, 'core/teacher/book_appointment.html', {
        'lang': lang,
        'students': students,
        'selected_student': selected_student,
        'selected_student_id': selected_student_id,
        'slots_by_date': slots_by_date,
        'reason_choices': reason_choices,
        'success': success,
        'error': error,
        'teacher_class': teacher_class,
    })


# ══════════════════════════════════════════════════════════════
# ПСИХОЛОГИЧЕСКИЙ ТЕСТ-ЦЕНТР
# ══════════════════════════════════════════════════════════════

@login_required
def psych_test_list(request):
    """Список тестов психолога"""
    lang = get_lang(request)
    if get_role(request.user) != 'psychologist': return role_redirect(request.user)
    from .models import PsychTest, SchoolClass
    tests = PsychTest.objects.filter(psychologist=request.user).prefetch_related('questions','results')
    all_classes = sorted(SchoolClass.objects.all(), key=lambda c: _class_sort_key(c.name))
    from .models import UserProfile, AnonymousRequest
    return render(request, 'core/psychologist/test_list.html', {
        'lang': lang, 'tests': tests, 'all_classes': all_classes,
        'student_count': UserProfile.objects.filter(role='student').count(),
        'class_count': all_classes.__len__(),
        'new_request_count': AnonymousRequest.objects.filter(status='new').count(),
    })


@login_required
def psych_test_create(request):
    """Создание теста — конструктор или ИИ"""
    lang = get_lang(request)
    if get_role(request.user) != 'psychologist': return role_redirect(request.user)
    from .models import PsychTest, PsychTestQuestion, SchoolClass
    import json as json_mod

    all_classes = sorted(SchoolClass.objects.all(), key=lambda c: _class_sort_key(c.name))
    source = request.GET.get('source', 'manual')  # 'manual' или 'ai'

    if request.method == 'POST':
        title      = request.POST.get('title','').strip()
        description= request.POST.get('description','').strip()
        topic      = request.POST.get('topic','anxiety')
        difficulty = request.POST.get('difficulty','medium')
        class_id   = request.POST.get('target_class','')
        source_post= request.POST.get('source','manual')

        target_class = None
        if class_id:
            try: target_class = SchoolClass.objects.get(id=class_id)
            except: pass

        test = PsychTest.objects.create(
            title=title, description=description,
            psychologist=request.user, source=source_post,
            topic=topic, difficulty=difficulty,
            target_class=target_class, status='draft'
        )

        if source_post == 'ai':
            # ── ИИ эмулятор — генерируем вопросы ──────────────────
            ai_questions = _generate_ai_questions(topic, difficulty)
            for i, q in enumerate(ai_questions):
                PsychTestQuestion.objects.create(
                    test=test, text=q['text'], q_type=q['type'],
                    options=q.get('options',[]), weight=q.get('weight',1), order=i
                )
            test.status = 'draft'
            test.save()
        else:
            # Конструктор — вопросы добавляются по одному
            questions_json = request.POST.get('questions_json','[]')
            try:
                questions = json_mod.loads(questions_json)
                for i, q in enumerate(questions):
                    PsychTestQuestion.objects.create(
                        test=test, text=q.get('text',''), q_type=q.get('type','radio'),
                        options=q.get('options',[]), weight=int(q.get('weight',1)), order=i
                    )
            except: pass

        return redirect('psych_test_detail', test_id=test.id)

    return render(request, 'core/psychologist/test_create.html', {
        'lang': lang, 'source': source, 'all_classes': all_classes,
        'topic_choices': PsychTest.TOPIC_CHOICES,
        'difficulty_choices': PsychTest.DIFFICULTY_CHOICES,
    })


def _generate_ai_questions(topic, difficulty):
    """ИИ эмулятор — возвращает набор вопросов по теме"""
    banks = {
        'anxiety': [
            {'text': 'Как часто вы чувствуете беспокойство без видимой причины?', 'type': 'scale', 'options': [], 'weight': 2},
            {'text': 'Вам трудно расслабиться даже когда всё хорошо?', 'type': 'radio', 'options': ['Никогда','Иногда','Часто','Всегда'], 'weight': 2},
            {'text': 'Возникают ли у вас мысли о плохом исходе событий?', 'type': 'radio', 'options': ['Никогда','Редко','Иногда','Часто'], 'weight': 2},
            {'text': 'Как вы обычно себя чувствуете перед важным событием?', 'type': 'radio', 'options': ['Спокойно','Немного волнуюсь','Очень волнуюсь','Паникую'], 'weight': 1},
            {'text': 'Есть ли у вас проблемы со сном из-за беспокойных мыслей?', 'type': 'radio', 'options': ['Никогда','Иногда','Часто','Каждую ночь'], 'weight': 2},
            {'text': 'Опишите своими словами что вас беспокоит больше всего:', 'type': 'text', 'options': [], 'weight': 1},
        ],
        'stress': [
            {'text': 'Как часто вы чувствуете, что у вас слишком много дел?', 'type': 'scale', 'options': [], 'weight': 2},
            {'text': 'Успеваете ли вы выполнять домашние задания вовремя?', 'type': 'radio', 'options': ['Всегда','Обычно','Иногда нет','Редко'], 'weight': 1},
            {'text': 'Как вы себя чувствуете в конце учебного дня?', 'type': 'radio', 'options': ['Бодро','Немного устал','Очень устал','Полностью истощён'], 'weight': 2},
            {'text': 'Бывают ли у вас головные боли или боли в животе от стресса?', 'type': 'radio', 'options': ['Никогда','Иногда','Часто','Постоянно'], 'weight': 2},
            {'text': 'Что помогает вам справляться со стрессом?', 'type': 'text', 'options': [], 'weight': 1},
        ],
        'selfesteem': [
            {'text': 'Как вы оцениваете себя по сравнению с одноклассниками?', 'type': 'scale', 'options': [], 'weight': 2},
            {'text': 'Вы довольны собой в целом?', 'type': 'radio', 'options': ['Очень доволен','Доволен','Не совсем','Совсем нет'], 'weight': 2},
            {'text': 'Боитесь ли вы осуждения со стороны других?', 'type': 'radio', 'options': ['Никогда','Иногда','Часто','Постоянно'], 'weight': 2},
            {'text': 'Какие качества в себе вам нравятся больше всего?', 'type': 'text', 'options': [], 'weight': 1},
            {'text': 'Умеете ли вы принимать критику?', 'type': 'radio', 'options': ['Легко','С трудом','Очень тяжело','Не умею'], 'weight': 1},
        ],
        'bullying': [
            {'text': 'Случалось ли, что одноклассники обижали вас словами или действиями?', 'type': 'radio', 'options': ['Никогда','Один раз','Иногда','Часто'], 'weight': 3},
            {'text': 'Чувствуете ли вы себя в безопасности в школе?', 'type': 'scale', 'options': [], 'weight': 3},
            {'text': 'Есть ли у вас близкие друзья в классе?', 'type': 'radio', 'options': ['Да, много','Есть несколько','Почти нет','Нет'], 'weight': 2},
            {'text': 'Рассказывали ли вы взрослым о проблемах в школе?', 'type': 'radio', 'options': ['Да, всегда','Иногда','Редко','Никогда'], 'weight': 2},
            {'text': 'Опишите ситуацию если вам некомфортно в коллективе:', 'type': 'text', 'options': [], 'weight': 1},
        ],
        'motivation': [
            {'text': 'Насколько вам интересна учёба?', 'type': 'scale', 'options': [], 'weight': 2},
            {'text': 'Есть ли предметы которые вам нравятся?', 'type': 'radio', 'options': ['Много','Несколько','Один-два','Нет'], 'weight': 1},
            {'text': 'Хотите ли вы продолжать образование после школы?', 'type': 'radio', 'options': ['Очень хочу','Наверное да','Не уверен','Нет'], 'weight': 2},
            {'text': 'Что мешает вам учиться лучше?', 'type': 'text', 'options': [], 'weight': 1},
            {'text': 'Как часто вы откладываете домашние задания?', 'type': 'radio', 'options': ['Никогда','Иногда','Часто','Всегда'], 'weight': 2},
        ],
    }
    questions = banks.get(topic, banks['anxiety'])
    if difficulty == 'easy':
        questions = questions[:4]
    elif difficulty == 'hard':
        extra = [
            {'text': 'Как вы справляетесь с трудными ситуациями?', 'type': 'text', 'options': [], 'weight': 1},
            {'text': 'Обращались ли вы за помощью к специалистам ранее?', 'type': 'radio', 'options': ['Нет','Один раз','Несколько раз','Регулярно'], 'weight': 2},
        ]
        questions = questions + extra
    return questions


@login_required
def psych_test_detail(request, test_id):
    """Детальный просмотр теста + назначение классу"""
    lang = get_lang(request)
    if get_role(request.user) != 'psychologist': return role_redirect(request.user)
    from .models import PsychTest, SchoolClass

    test = get_object_or_404(PsychTest, id=test_id, psychologist=request.user)
    all_classes = sorted(SchoolClass.objects.all(), key=lambda c: _class_sort_key(c.name))
    results = test.results.select_related('student').order_by('-completed_at')

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'activate':
            class_id = request.POST.get('target_class','')
            if class_id:
                try:
                    test.target_class = SchoolClass.objects.get(id=class_id)
                except: pass
            test.status = 'active'
            test.save()
            # Уведомляем учеников класса
            if test.target_class:
                students = User.objects.filter(
                    profile__role='student',
                    profile__school_class=test.target_class
                )
                for student in students:
                    create_notification(
                        student, 'appointment_new',
                        f'Психолог назначил вам тест: «{test.title}». Пройдите его в Тест-центре!',
                        '/student/psych-tests/'
                    )
        elif action == 'close':
            test.status = 'closed'
            test.save()
        elif action == 'draft':
            test.status = 'draft'
            test.save()
        return redirect('psych_test_detail', test_id=test.id)

    return render(request, 'core/psychologist/test_detail.html', {
        'lang': lang, 'test': test, 'results': results, 'all_classes': all_classes,
    })


@login_required
def psych_test_result_detail(request, result_id):
    """Результат конкретного ученика + заключение психолога"""
    lang = get_lang(request)
    if get_role(request.user) != 'psychologist': return role_redirect(request.user)
    from .models import PsychTestResult

    result = get_object_or_404(PsychTestResult, id=result_id, test__psychologist=request.user)

    if request.method == 'POST':
        conclusion = request.POST.get('conclusion','').strip()
        result.psy_conclusion = conclusion
        result.is_approved = True
        result.save()
        return redirect('psych_test_result_detail', result_id=result.id)

    # Готовим данные для графика
    questions = result.test.questions.all()
    chart_data = []
    for q in questions:
        ans = result.answers.get(str(q.id), '')
        if q.q_type == 'scale':
            try: val = int(ans)
            except: val = 0
            chart_data.append({'label': q.text[:30]+'...', 'value': val, 'max': 5})
        elif q.q_type in ['radio','checkbox']:
            opts = q.options
            if opts and ans in opts:
                val = opts.index(ans) + 1
                chart_data.append({'label': q.text[:30]+'...', 'value': val, 'max': len(opts)})

    # Готовим ответы для отображения
    answers_display = []
    for q in questions:
        ans = result.answers.get(str(q.id), '—')
        answers_display.append({'question': q, 'answer': ans})

    return render(request, 'core/psychologist/test_result_detail.html', {
        'lang': lang, 'result': result, 'questions': questions,
        'chart_data': json.dumps(chart_data),
        'answers_display': answers_display,
    })


@login_required
def psych_test_delete(request, test_id):
    """Удалить тест"""
    if get_role(request.user) != 'psychologist': return role_redirect(request.user)
    from .models import PsychTest
    test = get_object_or_404(PsychTest, id=test_id, psychologist=request.user)
    if request.method == 'POST':
        test.delete()
    return redirect('psych_test_list')


# ── УЧЕНИК — прохождение теста ────────────────────────────────

@login_required
def student_psych_tests(request):
    """Список доступных тестов для ученика"""
    lang = get_lang(request)
    if get_role(request.user) != 'student': return role_redirect(request.user)
    from .models import PsychTest, PsychTestResult

    student_class = getattr(request.user.profile, 'school_class', None)
    available_tests = PsychTest.objects.filter(
        status='active',
        target_class=student_class
    ) if student_class else PsychTest.objects.none()

    # Уже пройденные
    completed_ids = PsychTestResult.objects.filter(
        student=request.user
    ).values_list('test_id', flat=True)

    return render(request, 'core/student/psych_tests.html', {
        'lang': lang,
        'available_tests': available_tests,
        'completed_ids': list(completed_ids),
    })


@login_required
def student_take_psych_test(request, test_id):
    """Прохождение теста учеником"""
    lang = get_lang(request)
    if get_role(request.user) != 'student': return role_redirect(request.user)
    from .models import PsychTest, PsychTestResult

    test = get_object_or_404(PsychTest, id=test_id, status='active')

    # Уже прошёл?
    if PsychTestResult.objects.filter(test=test, student=request.user).exists():
        return redirect('student_psych_tests')

    questions = test.questions.all()

    if request.method == 'POST':
        answers = {}
        score = 0
        max_score = 0
        for q in questions:
            key = f'q_{q.id}'
            ans = request.POST.get(key, '')
            answers[str(q.id)] = ans
            # Подсчёт баллов
            if q.q_type == 'scale':
                try:
                    val = int(ans)
                    score += val * q.weight
                    max_score += 5 * q.weight
                except: max_score += 5 * q.weight
            elif q.q_type in ['radio','checkbox']:
                if q.options and ans in q.options:
                    val = q.options.index(ans) + 1
                    score += val * q.weight
                max_score += len(q.options) * q.weight if q.options else 0

        # Группа риска
        pct = (score / max_score * 100) if max_score > 0 else 0
        if pct >= 70: risk = 'high'
        elif pct >= 40: risk = 'medium'
        else: risk = 'low'

        # ИИ анализ — эмулятор
        ai_analysis = _generate_ai_analysis(test.topic, pct, risk)

        PsychTestResult.objects.create(
            test=test, student=request.user,
            answers=answers, score=score, max_score=max_score,
            risk_level=risk, ai_analysis=ai_analysis
        )

        # Уведомление психологу
        create_notification(
            test.psychologist, 'appointment_new',
            f'{request.user.get_full_name() or request.user.username} прошёл тест «{test.title}»',
            f'/psychologist/tests/{test.id}/'
        )

        return redirect('student_psych_test_done', test_id=test.id)

    return render(request, 'core/student/take_psych_test.html', {
        'lang': lang, 'test': test, 'questions': questions,
    })


def _generate_ai_analysis(topic, percent, risk):
    """ИИ эмулятор — генерирует текстовый анализ"""
    templates = {
        'high': {
            'anxiety': 'Результаты теста указывают на высокий уровень тревожности. Ученик испытывает значительный эмоциональный дискомфорт, который может влиять на учёбу и социальные отношения. Рекомендуется индивидуальная консультация с психологом.',
            'stress': 'Зафиксирован высокий уровень стресса. Ученик находится в состоянии хронического перегрузки. Необходимо снизить нагрузку и проработать стратегии совладания со стрессом.',
            'selfesteem': 'Выявлен низкий уровень самооценки. Ученик испытывает трудности с принятием себя, что может приводить к избеганию социальных ситуаций. Рекомендуется работа по формированию позитивного образа "я".',
            'bullying': 'Результаты свидетельствуют о возможных проблемах с безопасностью в коллективе. Требуется немедленное внимание — индивидуальная беседа с учеником и мониторинг ситуации в классе.',
            'motivation': 'Выявлен критически низкий уровень учебной мотивации. Ученик нуждается в поиске внутренних ресурсов и поддержке в определении личных целей.',
        },
        'medium': {
            'anxiety': 'Умеренный уровень тревожности в пределах нормы, однако требует наблюдения. Рекомендуется повторное тестирование через 2-3 недели.',
            'stress': 'Средний уровень стресса. Ученик справляется, но иногда испытывает перегрузку. Рекомендуется беседа о способах управления нагрузкой.',
            'selfesteem': 'Самооценка на среднем уровне. Есть зоны неуверенности, которые поддаются коррекции через групповые занятия.',
            'bullying': 'Есть отдельные признаки дискомфорта в коллективе. Рекомендуется профилактическая беседа.',
            'motivation': 'Мотивация неустойчива. Ученик может вдохновляться, но быстро теряет интерес. Нужна работа с целеполаганием.',
        },
        'low': {
            'anxiety': 'Уровень тревожности в норме. Ученик демонстрирует хорошую эмоциональную устойчивость. Профилактические меры не требуются.',
            'stress': 'Стрессоустойчивость на высоком уровне. Ученик эффективно справляется с нагрузками.',
            'selfesteem': 'Самооценка адекватная и стабильная. Ученик уверен в себе и позитивно воспринимает окружающих.',
            'bullying': 'Ученик чувствует себя в безопасности в коллективе, имеет дружеские связи.',
            'motivation': 'Высокий уровень учебной мотивации. Ученик целеустремлён и заинтересован в обучении.',
        }
    }
    text = templates.get(risk, {}).get(topic, f'Результат теста: {percent:.0f}%. Уровень риска: {risk}.')
    return f"[ИИ Анализ — эмулятор]\n\nБалл: {percent:.0f}%\nГруппа риска: {'Высокий' if risk=='high' else 'Средний' if risk=='medium' else 'Низкий'}\n\n{text}"


@login_required
def student_psych_test_done(request, test_id):
    """Страница после прохождения теста"""
    lang = get_lang(request)
    from .models import PsychTest, PsychTestResult
    test = get_object_or_404(PsychTest, id=test_id)
    result = get_object_or_404(PsychTestResult, test=test, student=request.user)
    return render(request, 'core/student/psych_test_done.html', {
        'lang': lang, 'test': test, 'result': result,
    })


# ══════════════════════════════════════════════════════════════
# ЦИКЛ ОБРАЩЕНИЙ — Шаг 1: Подача жалобы (ученик)
# ══════════════════════════════════════════════════════════════

@login_required
def student_new_request(request):
    lang = get_lang(request)
    if get_role(request.user) != 'student':
        return role_redirect(request.user)

    error = None

    if request.method == 'POST':
        text        = request.POST.get('text', '').strip()
        category    = request.POST.get('category', 'other')
        is_anon     = request.POST.get('is_anonymous') == 'on'

        if not text:
            error = ('Жағдайды сипаттаңыз' if lang == 'kz'
                     else 'Опишите ситуацию — это поможет психологу быстрее понять вас')
        else:
            req = StudentRequest.objects.create(
                student=request.user,
                category=category,
                text=text,
                is_anonymous=is_anon,
                status='new',
                lang=lang,
            )

            # Шаг 2: сразу запускаем ИИ-анализ
            _analyze_complaint(req)

            display_name = ('***' if is_anon
                            else (request.user.get_full_name() or request.user.username))
            risk_label = {'low': '🟢', 'medium': '🟡', 'high': '🔴', 'critical': '🚨'}
            risk_icon = risk_label.get(req.ai_risk_level, '')

            psychologists = User.objects.filter(profile__role='psychologist')
            for psy in psychologists:
                create_notification(
                    psy, 'request_ai_ready',
                    f'{risk_icon} Новое обращение [{req.get_ai_risk_level_display()}]: '
                    f'{display_name} — {req.get_category_display()}',
                    f'/psychologist/requests/{req.id}/',
                )

            return redirect(f'{request.path}?sent=1')

    sent = request.GET.get('sent') == '1'
    past_requests = StudentRequest.objects.filter(
        student=request.user
    ).order_by('-created_at')[:5]

    return render(request, 'core/student/new_request.html', {
        'lang': lang,
        'category_choices': StudentRequest.CATEGORY_CHOICES,
        'error': error,
        'sent': sent,
        'past_requests': past_requests,
        'active': 'requests',
    })


# ══════════════════════════════════════════════════════════════
# ЦИКЛ ОБРАЩЕНИЙ — Шаг 1: Поведенческий сигнал (учитель)
# ══════════════════════════════════════════════════════════════

@login_required
def teacher_new_observation(request):
    lang = get_lang(request)
    if get_role(request.user) != 'teacher':
        return role_redirect(request.user)

    teacher_class = getattr(request.user.profile, 'school_class', None)

    if teacher_class:
        students = User.objects.filter(
            profile__role='student',
            profile__school_class=teacher_class,
        ).order_by('last_name', 'first_name')
    else:
        students = User.objects.filter(
            profile__role='student',
        ).order_by('last_name', 'first_name')

    error = None

    if request.method == 'POST':
        student_id   = request.POST.get('student_id', '').strip()
        category     = request.POST.get('category', 'other')
        urgency      = request.POST.get('urgency', 'low')
        description  = request.POST.get('description', '').strip()
        duration_raw = request.POST.get('duration_days', '').strip()

        if not student_id or not description:
            error = ('Міндетті өрістерді толтырыңыз' if lang == 'kz'
                     else 'Заполните обязательные поля: ученик и описание')
        else:
            try:
                student = User.objects.get(id=student_id, profile__role='student')

                duration = int(duration_raw) if duration_raw.isdigit() else None

                obs = TeacherObservation.objects.create(
                    teacher=request.user,
                    student=student,
                    category=category,
                    urgency=urgency,
                    description=description,
                    duration_days=duration,
                )

                teacher_name = request.user.get_full_name() or request.user.username
                student_name = student.get_full_name() or student.username
                urgency_label = dict(TeacherObservation.URGENCY_CHOICES).get(urgency, urgency)

                psychologists = User.objects.filter(profile__role='psychologist')
                for psy in psychologists:
                    create_notification(
                        psy, 'observation_new',
                        f'Сигнал от учителя {teacher_name}: {student_name} — '
                        f'{obs.get_category_display()} [{urgency_label}]',
                        f'/psychologist/observations/{obs.id}/',
                    )

                return redirect(f'{request.path}?sent=1')

            except User.DoesNotExist:
                error = ('Оқушы табылмады' if lang == 'kz' else 'Ученик не найден')

    sent = request.GET.get('sent') == '1'
    recent_obs = TeacherObservation.objects.filter(
        teacher=request.user
    ).select_related('student').order_by('-created_at')[:8]

    return render(request, 'core/teacher/new_observation.html', {
        'lang': lang,
        'students': students,
        'teacher_class': teacher_class,
        'category_choices': TeacherObservation.CATEGORY_CHOICES,
        'urgency_choices': TeacherObservation.URGENCY_CHOICES,
        'error': error,
        'sent': sent,
        'recent_obs': recent_obs,
        'active': 'observe',
    })


# ══════════════════════════════════════════════════════════════
# ЦИКЛ ОБРАЩЕНИЙ — Шаг 3: Кейсы психолога
# ══════════════════════════════════════════════════════════════

_RISK_SORT = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3, '': 4}


@login_required
def psychologist_requests_list(request):
    lang = get_lang(request)
    if get_role(request.user) != 'psychologist':
        return role_redirect(request.user)

    status_f = request.GET.get('status', '')
    risk_f   = request.GET.get('risk', '')

    qs = StudentRequest.objects.select_related(
        'student', 'student__profile__school_class', 'assigned_to'
    ).exclude(status='archived')

    if status_f:
        qs = qs.filter(status=status_f)
    if risk_f:
        # effective_risk is a property — filter on both AI and override
        qs = qs.filter(
            models.Q(psy_risk_override=risk_f) |
            models.Q(psy_risk_override='', ai_risk_level=risk_f)
        )

    requests_list = sorted(
        qs,
        key=lambda r: (_RISK_SORT.get(r.effective_risk, 4), -r.created_at.timestamp())
    )

    observations_list = TeacherObservation.objects.select_related(
        'student', 'student__profile__school_class', 'teacher'
    ).filter(is_reviewed=False).order_by(
        models.Case(
            models.When(urgency='critical', then=0),
            models.When(urgency='high', then=1),
            models.When(urgency='medium', then=2),
            default=3,
            output_field=models.IntegerField(),
        ),
        '-created_at'
    )
    new_obs_count = observations_list.count()

    stats = {
        'new':       StudentRequest.objects.filter(status__in=['new', 'ai_analyzed']).count(),
        'in_review': StudentRequest.objects.filter(status='in_review').count(),
        'high_risk': StudentRequest.objects.filter(
            ai_risk_level__in=['high', 'critical']
        ).exclude(status__in=['concluded', 'archived']).count(),
        'total':     StudentRequest.objects.exclude(status='archived').count(),
    }

    return render(request, 'core/psychologist/requests_list.html', {
        'lang': lang,
        'requests_list': requests_list,
        'status_f': status_f,
        'risk_f': risk_f,
        'new_obs_count': new_obs_count,
        'observations_list': observations_list,
        'stats': stats,
        'status_choices': StudentRequest.STATUS_CHOICES,
        'risk_choices': StudentRequest.RISK_CHOICES,
        'active': 'cases',
    })


@login_required
def psychologist_request_detail(request, req_id):
    lang = get_lang(request)
    if get_role(request.user) != 'psychologist':
        return role_redirect(request.user)

    from .models import ParentStudent
    req = get_object_or_404(StudentRequest, id=req_id)
    student = req.student

    # Переводим в in_review при первом открытии
    if req.status in ('new', 'ai_analyzed'):
        req.status = 'in_review'
        if not req.assigned_to:
            req.assigned_to = request.user
        req.save(update_fields=['status', 'assigned_to'])

    # Помечаем наблюдения учителя просмотренными
    TeacherObservation.objects.filter(
        student=student, is_reviewed=False
    ).update(is_reviewed=True, reviewed_by=request.user, reviewed_at=timezone.now())

    # Контекст ученика
    observations  = TeacherObservation.objects.filter(
        student=student
    ).select_related('teacher').order_by('-created_at')[:10]

    recent_emotions = EmotionEntry.objects.filter(
        user=student
    ).order_by('-created_at')[:7]

    past_requests = StudentRequest.objects.filter(
        student=student
    ).exclude(id=req.id).order_by('-created_at')[:5]

    last_psych = PsychTestResult.objects.filter(
        student=student
    ).order_by('-completed_at').first()

    if request.method == 'POST':
        conclusion   = request.POST.get('psy_conclusion', '').strip()
        recs         = request.POST.get('psy_recommendations', '').strip()
        risk_override = request.POST.get('psy_risk_override', '').strip()
        parent_msg   = request.POST.get('parent_message', '').strip()
        do_approve   = request.POST.get('approve') == '1'

        req.psy_conclusion      = conclusion
        req.psy_recommendations = recs
        req.psy_risk_override   = risk_override
        req.parent_message      = parent_msg

        if do_approve and conclusion:
            req.is_approved  = True
            req.approved_at  = timezone.now()
            req.status       = 'concluded'

            parent_links = ParentStudent.objects.filter(
                student=student
            ).select_related('parent')
            for link in parent_links:
                create_notification(
                    link.parent, 'parent_conclusion',
                    f'Психолог подготовил заключение по обращению вашего ребёнка',
                    f'/parent/conclusions/{req.id}/',
                )
            req.parent_notified = bool(parent_links)
            if req.parent_notified:
                req.parent_notified_at = timezone.now()

        req.save()
        return redirect('psychologist_request_detail', req_id=req.id)

    return render(request, 'core/psychologist/request_detail.html', {
        'lang': lang,
        'req': req,
        'student': student,
        'observations': observations,
        'recent_emotions': recent_emotions,
        'past_requests': past_requests,
        'last_psych': last_psych,
        'ai_lines': req.ai_summary.split('\n') if req.ai_summary else [],
        'risk_choices': StudentRequest.RISK_CHOICES,
        'active': 'cases',
    })


@login_required
def psychologist_observation_detail(request, obs_id):
    lang = get_lang(request)
    if get_role(request.user) != 'psychologist':
        return role_redirect(request.user)

    obs = get_object_or_404(TeacherObservation, id=obs_id)

    if not obs.is_reviewed:
        obs.is_reviewed  = True
        obs.reviewed_by  = request.user
        obs.reviewed_at  = timezone.now()
        obs.save(update_fields=['is_reviewed', 'reviewed_by', 'reviewed_at'])

    if request.method == 'POST':
        obs.psy_note = request.POST.get('psy_note', '').strip()
        link_id = request.POST.get('linked_request_id', '').strip()
        if link_id:
            try:
                obs.linked_request = StudentRequest.objects.get(
                    id=link_id, student=obs.student
                )
            except StudentRequest.DoesNotExist:
                pass
        obs.save()
        return redirect('psychologist_observation_detail', obs_id=obs.id)

    student_requests = StudentRequest.objects.filter(
        student=obs.student
    ).order_by('-created_at')[:10]

    return render(request, 'core/psychologist/observation_detail.html', {
        'lang': lang,
        'obs': obs,
        'student_requests': student_requests,
        'active': 'cases',
    })


# ══════════════════════════════════════════════════════════════
# ЦИКЛ ОБРАЩЕНИЙ — Шаг 4: Заключение для родителя
# ══════════════════════════════════════════════════════════════

@login_required
def parent_conclusion_view(request, req_id):
    lang = get_lang(request)
    if get_role(request.user) != 'parent':
        return role_redirect(request.user)

    from .models import ParentStudent

    # Доступ только если ребёнок привязан к этому родителю
    req = get_object_or_404(
        StudentRequest,
        id=req_id,
        is_approved=True,
        student__parent_links__parent=request.user,
    )

    # Помечаем уведомления о заключении как прочитанные
    Notification.objects.filter(
        user=request.user,
        notif_type='parent_conclusion',
        link__contains=f'/parent/conclusions/{req.id}/',
        is_read=False,
    ).update(is_read=True)

    # Другие заключения по этому же ребёнку
    other_conclusions = StudentRequest.objects.filter(
        student=req.student,
        is_approved=True,
    ).exclude(id=req.id).order_by('-approved_at')[:4]

    return render(request, 'core/parent/conclusion_detail.html', {
        'lang': lang,
        'req': req,
        'other_conclusions': other_conclusions,
        'active': 'conclusions',
    })
